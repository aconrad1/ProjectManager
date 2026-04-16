"""Tests for Phase 4 — Gantt unscheduled section, right-click actions,
   enhanced Timeline integrity detection, and formula-target validation.
"""

import sys
import pytest
from pathlib import Path
from datetime import date, timedelta
from unittest.mock import MagicMock, patch, PropertyMock

from openpyxl import Workbook

# Ensure scripts/ is on sys.path so that `from gui.base_page import ...` works
# when importing gantt_page from tests.
_SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from helpers.schema.gantt import (
    build_gantt_sheet, _classify_items, _date_range,
    FILL_SECTION, SECTION_FONT,
)
from helpers.schema.integrity import (
    check_timelines,
    check_and_repair,
    IntegrityReport,
    IntegrityIssue,
    _check_formula_targets,
    _check_duration_formulas,
    _VLOOKUP_SHEET_RE,
)
from helpers.schema.sheets import (
    SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES,
    SHEET_TIMELINES, SHEET_GANTT,
)
from helpers.schema.columns import (
    TIMELINES_COLUMNS, PROJECTS_COLUMNS, TASKS_COLUMNS,
    DELIVERABLES_COLUMNS, column_index as ci,
)
from helpers.data.tasks import clean


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _make_wb_with_data(
    projects=None, tasks=None, deliverables=None,
) -> Workbook:
    """Create a workbook with Projects/Tasks/Deliverables sheets.

    Each item is a dict: {"id": "...", "start": date | None, ...}.
    Columns are written in canonical order from the schema.
    """
    wb = Workbook()
    wb.active.title = "Overview"

    # ── Projects sheet ────────────────────────────────────────────────
    ws_p = wb.create_sheet(SHEET_PROJECTS)
    for i, col in enumerate(PROJECTS_COLUMNS, 1):
        ws_p.cell(row=1, column=i, value=col.name)
    for r, p in enumerate(projects or [], start=2):
        ws_p.cell(row=r, column=1, value=p["id"])
        ws_p.cell(row=r, column=ci(SHEET_PROJECTS, "Title") + 1,
                  value=p.get("title", ""))
        start_col = ci(SHEET_PROJECTS, "Start Date") + 1
        ws_p.cell(row=r, column=start_col, value=p.get("start"))

    # ── Tasks sheet ───────────────────────────────────────────────────
    ws_t = wb.create_sheet(SHEET_TASKS)
    for i, col in enumerate(TASKS_COLUMNS, 1):
        ws_t.cell(row=1, column=i, value=col.name)
    for r, t in enumerate(tasks or [], start=2):
        ws_t.cell(row=r, column=1, value=t["id"])
        ws_t.cell(row=r, column=ci(SHEET_TASKS, "Title") + 1,
                  value=t.get("title", ""))
        ws_t.cell(row=r, column=ci(SHEET_TASKS, "Project ID") + 1,
                  value=t.get("project_id", ""))
        start_col = ci(SHEET_TASKS, "Start Date") + 1
        ws_t.cell(row=r, column=start_col, value=t.get("start"))
        end_col = ci(SHEET_TASKS, "End Date") + 1
        ws_t.cell(row=r, column=end_col, value=t.get("end"))

    # ── Deliverables sheet ────────────────────────────────────────────
    ws_d = wb.create_sheet(SHEET_DELIVERABLES)
    for i, col in enumerate(DELIVERABLES_COLUMNS, 1):
        ws_d.cell(row=1, column=i, value=col.name)
    for r, d in enumerate(deliverables or [], start=2):
        ws_d.cell(row=r, column=1, value=d["id"])
        ws_d.cell(row=r, column=ci(SHEET_DELIVERABLES, "Title") + 1,
                  value=d.get("title", ""))
        ws_d.cell(row=r, column=ci(SHEET_DELIVERABLES, "Task ID") + 1,
                  value=d.get("task_id", ""))
        start_col = ci(SHEET_DELIVERABLES, "Start Date") + 1
        ws_d.cell(row=r, column=start_col, value=d.get("start"))

    return wb


def _make_timelines_sheet(wb: Workbook, items: list[dict]) -> None:
    """Manually construct a Timelines sheet for testing.

    Each item dict: {"id": "...", "type": "Task", formulas as strings, ...}
    """
    ws = wb.create_sheet(SHEET_TIMELINES) if SHEET_TIMELINES not in wb.sheetnames else wb[SHEET_TIMELINES]
    # Headers
    for i, col in enumerate(TIMELINES_COLUMNS, 1):
        ws.cell(row=1, column=i, value=col.name)
    # Data rows
    for r, item in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=item.get("id", ""))
        ws.cell(row=r, column=2, value=item.get("type", "Task"))
        ws.cell(row=r, column=3, value=item.get("title_formula",
                f'=IFERROR(VLOOKUP($A{r},\'Tasks\'!$A:$N,3,FALSE),"")'))
        ws.cell(row=r, column=4, value=item.get("parent_id", ""))
        ws.cell(row=r, column=5, value=item.get("start_formula",
                f'=IFERROR(VLOOKUP($A{r},\'Tasks\'!$A:$N,9,FALSE),"")'))
        ws.cell(row=r, column=6, value=item.get("duration_formula",
                f'=IFERROR(IF(F{r}="","",G{r}-F{r}),"")'))
        ws.cell(row=r, column=7, value=item.get("end_formula",
                f'=IFERROR(VLOOKUP($A{r},\'Tasks\'!$A:$N,10,FALSE),"")'))
        ws.cell(row=r, column=8, value=item.get("deadline_formula",
                f'=IFERROR(VLOOKUP($A{r},\'Tasks\'!$A:$N,11,FALSE),"")'))
        ws.cell(row=r, column=9, value=item.get("status_formula",
                f'=IFERROR(VLOOKUP($A{r},\'Tasks\'!$A:$N,7,FALSE),"")'))


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: _date_range helper
# ═══════════════════════════════════════════════════════════════════════════════

class TestDateRange:
    def test_daily(self):
        dates = _date_range(date(2026, 1, 1), date(2026, 1, 5), "daily")
        assert len(dates) == 5
        assert dates[0] == date(2026, 1, 1)
        assert dates[-1] == date(2026, 1, 5)

    def test_weekly(self):
        dates = _date_range(date(2026, 1, 1), date(2026, 2, 1), "weekly")
        assert all(
            (dates[i + 1] - dates[i]).days == 7
            for i in range(len(dates) - 1)
        )


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: _classify_items — Excel Gantt item splitting
# ═══════════════════════════════════════════════════════════════════════════════

class TestClassifyItems:
    def test_all_scheduled(self):
        """Tasks with start dates → all in scheduled list."""
        wb = _make_wb_with_data(
            tasks=[
                {"id": "T-001", "title": "A", "start": date(2026, 3, 1)},
                {"id": "T-002", "title": "B", "start": date(2026, 3, 5)},
            ],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        tl_ws = wb[SHEET_TIMELINES]
        scheduled, unscheduled = _classify_items(wb, tl_ws)
        assert "T-001" in scheduled
        assert "T-002" in scheduled
        assert len(unscheduled) == 0

    def test_mixed_scheduled_unscheduled(self):
        """Tasks without start dates → appear in unscheduled list."""
        wb = _make_wb_with_data(
            tasks=[
                {"id": "T-001", "title": "Dated", "start": date(2026, 3, 1)},
                {"id": "T-002", "title": "Undated", "start": None},
            ],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        tl_ws = wb[SHEET_TIMELINES]
        scheduled, unscheduled = _classify_items(wb, tl_ws)
        assert "T-001" in scheduled
        assert "T-002" in unscheduled

    def test_no_timelines_sheet(self):
        """No Timelines sheet → both lists empty."""
        scheduled, unscheduled = _classify_items(Workbook(), None)
        assert scheduled == []
        assert unscheduled == []

    def test_projects_and_deliverables(self):
        """Projects and deliverables also classified correctly."""
        wb = _make_wb_with_data(
            projects=[
                {"id": "P-001", "title": "Proj", "start": date(2026, 1, 1)},
                {"id": "P-002", "title": "NoStart", "start": None},
            ],
            deliverables=[
                {"id": "D-001", "title": "Del", "start": date(2026, 2, 1),
                 "task_id": "T-001"},
            ],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        tl_ws = wb[SHEET_TIMELINES]
        scheduled, unscheduled = _classify_items(wb, tl_ws)
        assert "P-001" in scheduled
        assert "P-002" in unscheduled
        assert "D-001" in scheduled


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: build_gantt_sheet — unscheduled section
# ═══════════════════════════════════════════════════════════════════════════════

class TestBuildGanttSheet:
    def test_unscheduled_section_present(self):
        """Gantt sheet includes 'No Scheduled Start' section header."""
        wb = _make_wb_with_data(
            tasks=[
                {"id": "T-001", "title": "Dated", "start": date(2026, 3, 1)},
                {"id": "T-002", "title": "Undated", "start": None},
            ],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        count = build_gantt_sheet(wb, start_date=date(2026, 3, 1))
        assert count == 2  # both items counted

        ws = wb[SHEET_GANTT]
        # Look for the section header
        section_found = False
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val == "No Scheduled Start":
                section_found = True
                break
        assert section_found

    def test_no_unscheduled_section_when_all_dated(self):
        """No section header when all items have start dates."""
        wb = _make_wb_with_data(
            tasks=[
                {"id": "T-001", "title": "A", "start": date(2026, 3, 1)},
            ],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        build_gantt_sheet(wb, start_date=date(2026, 3, 1))
        ws = wb[SHEET_GANTT]
        for r in range(2, ws.max_row + 1):
            assert ws.cell(row=r, column=1).value != "No Scheduled Start"

    def test_empty_gantt_returns_zero(self):
        """No items → return 0."""
        wb = _make_wb_with_data()
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        count = build_gantt_sheet(wb, start_date=date(2026, 1, 1))
        assert count == 0

    def test_unscheduled_items_have_vlookup(self):
        """Unscheduled items still get VLOOKUP formulas for title & status."""
        wb = _make_wb_with_data(
            tasks=[
                {"id": "T-001", "title": "Undated", "start": None},
            ],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        build_gantt_sheet(wb, start_date=date(2026, 3, 1))
        ws = wb[SHEET_GANTT]
        # Find the item row (after section header)
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "T-001":
                title_val = ws.cell(row=r, column=2).value
                assert title_val is not None
                assert "VLOOKUP" in str(title_val)
                break
        else:
            pytest.fail("T-001 not found in Gantt sheet")


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: VLOOKUP sheet-name regex
# ═══════════════════════════════════════════════════════════════════════════════

class TestVlookupSheetRegex:
    def test_extracts_sheet_name(self):
        formula = "=IFERROR(VLOOKUP($A2,'Tasks'!$A:$N,3,FALSE),\"\")"
        m = _VLOOKUP_SHEET_RE.search(formula)
        assert m is not None
        assert m.group(1) == "Tasks"

    def test_no_match_on_non_vlookup(self):
        formula = "=IF(A2>0,A2,0)"
        m = _VLOOKUP_SHEET_RE.search(formula)
        assert m is None


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: _check_formula_targets
# ═══════════════════════════════════════════════════════════════════════════════

class TestCheckFormulaTargets:
    def test_correct_targets_produce_no_issues(self):
        """VLOOKUP that references the right sheet → no issues."""
        wb = Workbook()
        ws = wb.create_sheet(SHEET_TIMELINES)
        for i, col in enumerate(TIMELINES_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col.name)
        ws.cell(row=2, column=1, value="T-001")
        ws.cell(row=2, column=3, value="=IFERROR(VLOOKUP($A2,'Tasks'!$A:$N,3,FALSE),\"\")")
        ws.cell(row=2, column=5, value="=IFERROR(VLOOKUP($A2,'Tasks'!$A:$N,9,FALSE),\"\")")
        ws.cell(row=2, column=7, value="=IFERROR(VLOOKUP($A2,'Tasks'!$A:$N,10,FALSE),\"\")")
        ws.cell(row=2, column=8, value="=IFERROR(VLOOKUP($A2,'Tasks'!$A:$N,11,FALSE),\"\")")
        ws.cell(row=2, column=9, value="=IFERROR(VLOOKUP($A2,'Tasks'!$A:$N,7,FALSE),\"\")")
        issues: list[IntegrityIssue] = []
        _check_formula_targets(ws, issues)
        assert len(issues) == 0

    def test_wrong_target_flagged(self):
        """VLOOKUP referencing 'Projects' for a T-xxx ID → error."""
        wb = Workbook()
        ws = wb.create_sheet(SHEET_TIMELINES)
        for i, col in enumerate(TIMELINES_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col.name)
        ws.cell(row=2, column=1, value="T-001")
        ws.cell(row=2, column=3,
                value="=IFERROR(VLOOKUP($A2,'Projects'!$A:$N,3,FALSE),\"\")")
        issues: list[IntegrityIssue] = []
        _check_formula_targets(ws, issues)
        assert len(issues) == 1
        assert issues[0].severity == "error"
        assert "Projects" in issues[0].message

    def test_project_vlookup_targets(self):
        """P-xxx items should reference the Projects sheet."""
        wb = Workbook()
        ws = wb.create_sheet(SHEET_TIMELINES)
        for i, col in enumerate(TIMELINES_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col.name)
        ws.cell(row=2, column=1, value="P-001")
        ws.cell(row=2, column=3,
                value="=IFERROR(VLOOKUP($A2,'Projects'!$A:$M,2,FALSE),\"\")")
        issues: list[IntegrityIssue] = []
        _check_formula_targets(ws, issues)
        assert len(issues) == 0  # correct sheet


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: _check_duration_formulas
# ═══════════════════════════════════════════════════════════════════════════════

class TestCheckDurationFormulas:
    def test_valid_duration_formula(self):
        wb = Workbook()
        ws = wb.create_sheet(SHEET_TIMELINES)
        for i, col in enumerate(TIMELINES_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col.name)
        ws.cell(row=2, column=1, value="T-001")
        ws.cell(row=2, column=6, value='=IFERROR(IF(F2="","",G2-F2),"")')
        issues: list[IntegrityIssue] = []
        _check_duration_formulas(ws, issues)
        assert len(issues) == 0

    def test_broken_duration_formula(self):
        wb = Workbook()
        ws = wb.create_sheet(SHEET_TIMELINES)
        for i, col in enumerate(TIMELINES_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col.name)
        ws.cell(row=2, column=1, value="T-001")
        ws.cell(row=2, column=6, value="=#REF!")
        issues: list[IntegrityIssue] = []
        _check_duration_formulas(ws, issues)
        assert len(issues) == 1
        assert issues[0].severity == "error"

    def test_empty_duration_is_warning(self):
        wb = Workbook()
        ws = wb.create_sheet(SHEET_TIMELINES)
        for i, col in enumerate(TIMELINES_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col.name)
        ws.cell(row=2, column=1, value="T-001")
        ws.cell(row=2, column=6, value=None)
        issues: list[IntegrityIssue] = []
        _check_duration_formulas(ws, issues)
        assert len(issues) == 1
        assert issues[0].severity == "warning"


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: check_timelines — enhanced checks
# ═══════════════════════════════════════════════════════════════════════════════

class TestCheckTimelinesEnhanced:
    def test_missing_sheet_is_error(self):
        wb = Workbook()
        report = check_timelines(wb)
        assert report.has_errors
        assert "missing" in report.issues[0].message.lower()

    def test_healthy_after_sync(self):
        """A freshly synced workbook should pass all checks."""
        wb = _make_wb_with_data(
            projects=[{"id": "P-001", "title": "P", "start": date(2026, 1, 1)}],
            tasks=[{"id": "T-001", "title": "T", "project_id": "P-001",
                    "start": date(2026, 1, 5)}],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        report = check_timelines(wb)
        assert not report.has_errors

    def test_detects_wrong_vlookup_target(self):
        """Hand-crafted wrong target → error in enhanced check."""
        wb = _make_wb_with_data(
            tasks=[{"id": "T-001", "title": "T", "start": date(2026, 1, 1)}],
        )
        _make_timelines_sheet(wb, [
            {"id": "T-001", "type": "Task",
             "title_formula": "=IFERROR(VLOOKUP($A2,'Projects'!$A:$N,3,FALSE),\"\")"},
        ])
        report = check_timelines(wb)
        assert report.has_errors
        target_issues = [i for i in report.issues if "references" in i.message]
        assert len(target_issues) >= 1


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: check_and_repair — rebuilds Gantt too
# ═══════════════════════════════════════════════════════════════════════════════

class TestCheckAndRepair:
    @patch("helpers.schema.gantt.build_gantt_sheet")
    @patch("helpers.schema.timelines.sync_timelines")
    def test_repair_rebuilds_gantt(self, mock_sync, mock_gantt):
        """When errors are detected, both Timelines and Gantt are rebuilt."""
        wb = Workbook()
        # First check_timelines returns errors, second returns clean
        with patch("helpers.schema.integrity.check_timelines") as mock_check:
            err_report = IntegrityReport(issues=[
                IntegrityIssue("error", None, None, "Timelines sheet is missing"),
            ])
            ok_report = IntegrityReport()
            mock_check.side_effect = [err_report, ok_report]

            report = check_and_repair(wb)
            assert report.repaired
            mock_sync.assert_called_once_with(wb)
            mock_gantt.assert_called_once_with(wb)

    def test_no_repair_when_healthy(self):
        """Healthy workbook → no rebuild."""
        wb = _make_wb_with_data(
            tasks=[{"id": "T-001", "title": "T", "start": date(2026, 1, 1)}],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)

        report = check_and_repair(wb)
        assert not report.repaired
        assert not report.has_errors

    def test_warnings_logged(self):
        """Warnings are reported but do not trigger repair."""
        wb = _make_wb_with_data(
            tasks=[{"id": "T-001", "title": "T", "start": date(2026, 1, 1)}],
        )
        from helpers.schema.timelines import sync_timelines
        sync_timelines(wb)
        # Inject a warning-only issue
        with patch("helpers.schema.integrity.check_timelines") as mock_check:
            warn_report = IntegrityReport(issues=[
                IntegrityIssue("warning", None, None, "Minor issue"),
            ])
            mock_check.return_value = warn_report

            logs = []
            report = check_and_repair(wb, log=logs.append)
            assert not report.repaired
            assert any("warning" in msg for msg in logs)


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: Gantt page — right-click helpers
# ═══════════════════════════════════════════════════════════════════════════════

class TestGanttPageRightClick:
    """Test the shift logic and row-hit detection (without spinning up tkinter)."""

    def test_shift_date_task(self):
        """_shift_date moves a task's start forward by 1 day."""
        from scripts.gui.pages.gantt_page import GanttPage

        mock_app = MagicMock()
        task = MagicMock()
        task.id = "T-001"
        task.start = date(2026, 4, 10)
        task.end = date(2026, 4, 20)
        mock_app.profile.find_by_id.return_value = task
        mock_app.service = MagicMock()

        page = object.__new__(GanttPage)
        page.app = mock_app
        page._rows = []
        page._row_y_ranges = []
        page._canvas = MagicMock()
        page._render = MagicMock()  # avoid tkinter widget access

        page._shift_date("T-001", "start", 1)

        mock_app.service.edit_task.assert_called_once_with(
            "T-001", {"start": date(2026, 4, 11)},
        )

    def test_shift_date_deliverable(self):
        """_shift_date moves a deliverable's end back by 1 day."""
        from scripts.gui.pages.gantt_page import GanttPage

        mock_app = MagicMock()
        deliv = MagicMock()
        deliv.id = "D-001"
        deliv.start = date(2026, 5, 1)
        deliv.end = date(2026, 5, 10)

        mock_app.profile.find_by_id.return_value = deliv
        mock_app.service = MagicMock()

        page = object.__new__(GanttPage)
        page.app = mock_app
        page._rows = []
        page._row_y_ranges = []
        page._canvas = MagicMock()
        page._render = MagicMock()  # avoid tkinter widget access

        page._shift_date("D-001", "end", -1)

        mock_app.service.edit_deliverable.assert_called_once_with(
            "D-001", {"end": date(2026, 5, 9)},
        )

    def test_shift_date_no_current_date(self):
        """Shifting a field that is None does nothing."""
        from scripts.gui.pages.gantt_page import GanttPage

        mock_app = MagicMock()
        task = MagicMock()
        task.id = "T-001"
        task.start = None
        mock_app.profile.find_by_id.return_value = task
        mock_app.service = MagicMock()

        page = object.__new__(GanttPage)
        page.app = mock_app
        page._rows = []
        page._row_y_ranges = []
        page._canvas = MagicMock()

        page._shift_date("T-001", "start", 1)

        mock_app.service.edit_task.assert_not_called()

    def test_hit_row_returns_correct_row(self):
        """_hit_row returns the row at the given y-coordinate."""
        from scripts.gui.pages.gantt_page import GanttPage
        from helpers.reporting.gantt import GanttRow

        page = object.__new__(GanttPage)
        page._rows = [
            GanttRow(type="project", label="Proj"),
            GanttRow(type="task", label="Task 1", item_id="T-001"),
        ]
        page._row_y_ranges = [(44, 70), (70, 96)]

        assert page._hit_row(50).type == "project"
        assert page._hit_row(75).item_id == "T-001"
        assert page._hit_row(100) is None


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: IntegrityReport properties
# ═══════════════════════════════════════════════════════════════════════════════

class TestIntegrityReportProperties:
    def test_healthy(self):
        r = IntegrityReport()
        assert r.is_healthy
        assert not r.has_errors
        assert not r.has_warnings
        assert r.error_count == 0
        assert r.warning_count == 0

    def test_mixed(self):
        r = IntegrityReport(issues=[
            IntegrityIssue("error", 2, "Title", "broken"),
            IntegrityIssue("warning", 3, "Duration (days)", "empty"),
            IntegrityIssue("error", 4, "Status", "broken"),
        ])
        assert r.has_errors
        assert r.has_warnings
        assert r.error_count == 2
        assert r.warning_count == 1
        assert not r.is_healthy


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST: Gantt page — unscheduled tasks appear in row data
# ═══════════════════════════════════════════════════════════════════════════════

class TestGanttPageRendering:
    """Validate the row-building logic produces unscheduled sections."""

    def _make_profile(self):
        """Build a mock profile with dated and undated tasks."""
        from helpers.domain.project import Project
        from helpers.domain.task import Task

        p = MagicMock()
        proj = MagicMock(spec=Project)
        proj.title = "TestProj"
        proj.category = "Ongoing"
        proj.tasks = []

        t1 = MagicMock(spec=Task)
        t1.id = "T-001"
        t1.title = "Dated Task"
        t1.start = date(2026, 3, 1)
        t1.end = date(2026, 3, 10)
        t1.deadline = None
        t1.status = "In Progress"
        t1.priority = 2
        t1.deliverables = []

        t2 = MagicMock(spec=Task)
        t2.id = "T-002"
        t2.title = "Undated Task"
        t2.start = None
        t2.end = None
        t2.deadline = None
        t2.status = "Not Started"
        t2.priority = 3
        t2.deliverables = []

        proj.tasks = [t1, t2]
        p.projects = [proj]
        return p

    def test_unscheduled_section_in_rows(self):
        """The _render logic should produce a 'section' row for undated tasks."""
        from scripts.gui.pages.gantt_page import GanttPage

        profile = self._make_profile()

        page = object.__new__(GanttPage)
        page.app = MagicMock()
        page.app.profile = profile
        page._filter_var = MagicMock()
        page._filter_var.get.return_value = "All"
        page._day_width_var = MagicMock()
        page._day_width_var.get.return_value = 16
        page._gantt_colors = {}
        page._gantt_colors_dark = {}
        page._dark_mode = False
        page._canvas = MagicMock()
        page._rows = []
        page._row_y_ranges = []

        # Call _render — the canvas calls are mocked, but _rows gets populated
        page._render()

        row_types = [r["type"] for r in page._rows]
        assert "section" in row_types
        section_idx = row_types.index("section")
        assert page._rows[section_idx]["label"] == "No Scheduled Start"
        # The undated task should follow the section header
        assert page._rows[section_idx + 1]["item_id"] == "T-002"
