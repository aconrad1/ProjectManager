"""Tests for Phase 3 — snapshot diffs, rolling windows, timelines integrity."""

import json
import pytest
from datetime import date, timedelta
from unittest.mock import MagicMock, patch, PropertyMock

from helpers.reporting.snapshot_diff import (
    diff_profiles,
    SnapshotDiff,
    EntityChange,
    FieldChange,
    _str,
    _diff_fields,
    baseline_profile_for_diff,
)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _make_deliverable(did="D-001", title="Del 1", status="Not Started",
                      percent_complete=0, time_allocated=1.0, time_spent=0.0):
    d = MagicMock()
    d.id = did
    d.title = title
    d.status = status
    d.percent_complete = percent_complete
    d.time_allocated = time_allocated
    d.time_spent = time_spent
    d.deliverables = []  # deliverables don't have sub-deliverables
    return d


def _make_task(tid="T-001", title="Task 1", status="In Progress",
               priority=3, supervisor="", site="", commentary="",
               deliverables=None):
    t = MagicMock()
    t.id = tid
    t.title = title
    t.status = status
    t.priority = priority
    t.supervisor = supervisor
    t.site = site
    t.commentary = commentary
    t.deliverables = deliverables or []
    return t


def _make_project(pid="P-001", title="Project 1", category="Ongoing",
                  status="In Progress", priority=3, supervisor="",
                  site="", tasks=None):
    p = MagicMock()
    p.id = pid
    p.title = title
    p.category = category
    p.status = status
    p.priority = priority
    p.supervisor = supervisor
    p.site = site
    p.tasks = tasks or []
    return p


def _make_profile(projects=None):
    profile = MagicMock()
    profile.projects = projects or []
    all_tasks = []
    for p in profile.projects:
        all_tasks.extend(p.tasks)
    profile.all_tasks = all_tasks
    return profile


# ══════════════════════════════════════════════════════════════════════════════
#  Snapshot Diff Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestStrHelper:
    def test_none(self):
        assert _str(None) == ""

    def test_date(self):
        assert _str(date(2026, 3, 25)) == "2026-03-25"

    def test_string_strip(self):
        assert _str("  hello  ") == "hello"

    def test_int(self):
        assert _str(3) == "3"


class TestDiffFields:
    def test_no_changes(self):
        obj = MagicMock(title="A", status="X")
        changes = _diff_fields(obj, obj, ("title", "status"))
        assert changes == []

    def test_detects_change(self):
        old = MagicMock(title="A", status="X")
        new = MagicMock(title="A", status="Y")
        changes = _diff_fields(old, new, ("title", "status"))
        assert len(changes) == 1
        assert changes[0].field == "status"
        assert changes[0].old == "X"
        assert changes[0].new == "Y"


class TestDiffProfiles:
    """Test the core diff_profiles function."""

    def test_no_changes(self):
        p1 = _make_project("P-001", "Proj", tasks=[
            _make_task("T-001", "Task 1")
        ])
        old = _make_profile([p1])
        new = _make_profile([p1])
        diff = diff_profiles(old, new)
        assert not diff.has_changes
        assert diff.changes == []

    def test_added_project(self):
        old = _make_profile([])
        p1 = _make_project("P-001", "New Project")
        new = _make_profile([p1])
        diff = diff_profiles(old, new)
        assert diff.has_changes
        assert len(diff.added) == 1
        assert diff.added[0].entity_type == "project"
        assert diff.added[0].entity_id == "P-001"

    def test_removed_project(self):
        p1 = _make_project("P-001", "Old Project")
        old = _make_profile([p1])
        new = _make_profile([])
        diff = diff_profiles(old, new)
        assert len(diff.removed) == 1
        assert diff.removed[0].entity_id == "P-001"

    def test_modified_project(self):
        p_old = _make_project("P-001", "Proj", status="In Progress")
        p_new = _make_project("P-001", "Proj", status="Completed")
        old = _make_profile([p_old])
        new = _make_profile([p_new])
        diff = diff_profiles(old, new)
        assert len(diff.modified) == 1
        assert diff.modified[0].fields[0].field == "status"

    def test_added_task(self):
        p_old = _make_project("P-001", "Proj", tasks=[])
        p_new = _make_project("P-001", "Proj", tasks=[
            _make_task("T-001", "New Task")
        ])
        old = _make_profile([p_old])
        new = _make_profile([p_new])
        diff = diff_profiles(old, new)
        assert len(diff.added) == 1
        assert diff.added[0].entity_type == "task"

    def test_removed_task(self):
        t = _make_task("T-001", "Old Task")
        p_old = _make_project("P-001", "Proj", tasks=[t])
        p_new = _make_project("P-001", "Proj", tasks=[])
        old = _make_profile([p_old])
        new = _make_profile([p_new])
        diff = diff_profiles(old, new)
        assert len(diff.removed) == 1
        assert diff.removed[0].entity_type == "task"

    def test_modified_task(self):
        t_old = _make_task("T-001", "Task", status="In Progress")
        t_new = _make_task("T-001", "Task", status="Completed")
        p_old = _make_project("P-001", "Proj", tasks=[t_old])
        p_new = _make_project("P-001", "Proj", tasks=[t_new])
        old = _make_profile([p_old])
        new = _make_profile([p_new])
        diff = diff_profiles(old, new)
        assert len(diff.modified) == 1
        assert diff.modified[0].fields[0].old == "In Progress"

    def test_added_deliverable(self):
        t_old = _make_task("T-001", "Task", deliverables=[])
        d_new = _make_deliverable("D-001", "New Del")
        t_new = _make_task("T-001", "Task", deliverables=[d_new])
        p_old = _make_project("P-001", "Proj", tasks=[t_old])
        p_new = _make_project("P-001", "Proj", tasks=[t_new])
        old = _make_profile([p_old])
        new = _make_profile([p_new])
        diff = diff_profiles(old, new)
        assert len(diff.added) == 1
        assert diff.added[0].entity_type == "deliverable"

    def test_modified_deliverable(self):
        d_old = _make_deliverable("D-001", "Del", percent_complete=0)
        d_new = _make_deliverable("D-001", "Del", percent_complete=50)
        t_old = _make_task("T-001", "Task", deliverables=[d_old])
        t_new = _make_task("T-001", "Task", deliverables=[d_new])
        p_old = _make_project("P-001", "Proj", tasks=[t_old])
        p_new = _make_project("P-001", "Proj", tasks=[t_new])
        old = _make_profile([p_old])
        new = _make_profile([p_new])
        diff = diff_profiles(old, new)
        assert len(diff.modified) == 1
        assert diff.modified[0].entity_type == "deliverable"

    def test_multiple_changes(self):
        """Test a mix of added, removed, and modified entities."""
        t1 = _make_task("T-001", "Same", status="In Progress")
        t2_old = _make_task("T-002", "Changed", status="In Progress")
        t2_new = _make_task("T-002", "Changed", status="Completed")
        t3 = _make_task("T-003", "Removed")
        t4 = _make_task("T-004", "Added")

        p_old = _make_project("P-001", "Proj", tasks=[t1, t2_old, t3])
        p_new = _make_project("P-001", "Proj", tasks=[t1, t2_new, t4])
        old = _make_profile([p_old])
        new = _make_profile([p_new])
        diff = diff_profiles(old, new)

        assert len(diff.added) == 1     # T-004
        assert len(diff.removed) == 1   # T-003
        assert len(diff.modified) == 1  # T-002 status


class TestSnapshotDiffProperties:
    def test_empty_diff(self):
        diff = SnapshotDiff()
        assert not diff.has_changes
        assert diff.added == []
        assert diff.removed == []
        assert diff.modified == []

    def test_summary_added(self):
        c = EntityChange("added", "task", "T-001", "New Task")
        assert "Added task" in c.summary

    def test_summary_removed(self):
        c = EntityChange("removed", "project", "P-001", "Old Proj")
        assert "Removed project" in c.summary

    def test_summary_modified(self):
        c = EntityChange("modified", "task", "T-001", "Task",
                         [FieldChange("status", "In Progress", "Completed")])
        assert "Modified task" in c.summary
        assert "status" in c.summary

    def test_baseline_profile_has_no_entities(self):
        t = _make_task("T-001", "Task")
        p = _make_project("P-001", "Project", tasks=[t])
        profile = _make_profile([p])
        profile.id = "profile-1"
        profile.title = "User"
        profile.description = "desc"
        profile.deadline = None
        profile.start = None
        profile.end = None
        profile.status = "Active"
        profile.company = "Co"
        profile.role = "Role"
        profile.email = "a@b.com"
        profile.phone = "123"
        profile.recipient_name = "Recipient"
        profile.recipient_email = "r@b.com"
        profile.workbook_filename = "w.xlsx"
        profile.daily_hours_budget = 8.0
        profile.weekly_hours_budget = 40.0

        baseline = baseline_profile_for_diff(profile)
        diff = diff_profiles(baseline, profile)

        assert baseline.projects == []
        assert diff.has_changes
        assert len(diff.added) == 2
        assert len(diff.removed) == 0
        assert len(diff.modified) == 0


# ══════════════════════════════════════════════════════════════════════════════
#  Rolling Windows Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestRollingWindows:
    """Test that markdown builder uses configurable windows from deadlines.json."""

    def _build_profile_with_completed(self, completed_date):
        """Build a minimal profile with one completed task."""
        d = _make_deliverable()
        t = _make_task("T-001", "Completed Task", status="Completed",
                       deliverables=[d])
        t.date_completed = completed_date
        t.site = "Test Site"
        t.supervisor = "Test Sup"
        t.commentary = ""
        p = _make_project("P-001", "Completed Project",
                          category="Completed", tasks=[t])

        profile = MagicMock()
        profile.projects = [p]
        profile.all_tasks = [t]

        def tasks_for_cat(cat):
            if cat.lower() == "completed":
                return [t]
            return []

        profile.tasks_for_category = tasks_for_cat
        return profile

    @patch("helpers.reporting.markdown.load_deadline_windows")
    def test_recent_window_default(self, mock_cfg):
        """With default config, uses 7-day window."""
        mock_cfg.return_value = {
            "recent_completed_days": 7,
            "extended_completed_days": 30,
        }
        from helpers.reporting.markdown import build_markdown

        today = date(2026, 3, 25)
        profile = self._build_profile_with_completed(today - timedelta(days=5))
        md = build_markdown(profile, [], today, author="Test", role="R", company="C")
        assert "Past 7 Days" in md

    @patch("helpers.reporting.markdown.load_deadline_windows")
    def test_extended_window_custom(self, mock_cfg):
        """Custom extended_completed_days = 14 should appear in heading."""
        mock_cfg.return_value = {
            "recent_completed_days": 7,
            "extended_completed_days": 14,
        }
        from helpers.reporting.markdown import build_markdown

        today = date(2026, 3, 25)
        profile = self._build_profile_with_completed(today - timedelta(days=10))
        md = build_markdown(profile, [], today, author="Test", role="R", company="C")
        assert "Last 14 Days" in md

    @patch("helpers.reporting.markdown.load_deadline_windows")
    def test_config_missing_fallback(self, mock_cfg):
        """If deadlines.json is missing, uses defaults."""
        mock_cfg.return_value = {
            "recent_completed_days": 7,
            "extended_completed_days": 30,
            "upcoming_deadline_days": 14,
            "snapshot_lookback_days": 7,
        }
        from helpers.reporting.markdown import build_markdown

        today = date(2026, 3, 25)
        profile = self._build_profile_with_completed(today - timedelta(days=5))
        md = build_markdown(profile, [], today, author="Test", role="R", company="C")
        # Defaults: 7 and 30
        assert "Past 7 Days" in md
        assert "Last 30 Days" in md


class TestDeadlineConfigRepair:
    def test_missing_deadlines_json_is_recreated(self, tmp_path, monkeypatch):
        from helpers.config import loader

        monkeypatch.setattr(loader, "_CONFIG_DIR", tmp_path)
        loader.load.cache_clear()

        logs: list[str] = []
        windows = loader.load_deadline_windows(log=logs.append)

        assert windows["recent_completed_days"] == 7
        assert (tmp_path / "deadlines.json").exists()
        assert any("missing" in msg.lower() for msg in logs)

    def test_invalid_deadlines_values_are_repaired(self, tmp_path, monkeypatch):
        from helpers.config import loader

        monkeypatch.setattr(loader, "_CONFIG_DIR", tmp_path)
        (tmp_path / "deadlines.json").write_text(json.dumps({
            "recent_completed_days": -3,
            "extended_completed_days": "abc",
            "upcoming_deadline_days": 10,
        }), encoding="utf-8")
        loader.load.cache_clear()

        logs: list[str] = []
        windows = loader.load_deadline_windows(log=logs.append)

        assert windows["recent_completed_days"] == 7
        assert windows["extended_completed_days"] == 30
        assert windows["upcoming_deadline_days"] == 10
        repaired = json.loads((tmp_path / "deadlines.json").read_text(encoding="utf-8"))
        assert repaired["snapshot_lookback_days"] == 7
        assert any("invalid" in msg.lower() for msg in logs)


# ══════════════════════════════════════════════════════════════════════════════
#  Change History in Markdown Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestChangeHistoryMarkdown:
    """Test that change history section appears in markdown output."""

    @patch("helpers.reporting.markdown.load_deadline_windows")
    def test_change_history_included(self, mock_cfg):
        mock_cfg.return_value = {
            "recent_completed_days": 7,
            "extended_completed_days": 30,
        }
        from helpers.reporting.markdown import build_markdown

        today = date(2026, 3, 25)
        profile = MagicMock()
        profile.tasks_for_category = lambda cat: []

        diff = SnapshotDiff(changes=[
            EntityChange("added", "task", "T-099", "Brand New Task"),
            EntityChange("modified", "project", "P-001", "Proj",
                         [FieldChange("status", "In Progress", "Completed")]),
        ])
        md = build_markdown(profile, [], today, author="A", role="R",
                            company="C", snapshot_diff=diff)
        assert "## Change History" in md
        assert "Brand New Task" in md
        assert "status" in md

    @patch("helpers.reporting.markdown.load_deadline_windows")
    def test_no_change_history_when_empty(self, mock_cfg):
        mock_cfg.return_value = {
            "recent_completed_days": 7,
            "extended_completed_days": 30,
        }
        from helpers.reporting.markdown import build_markdown

        today = date(2026, 3, 25)
        profile = MagicMock()
        profile.tasks_for_category = lambda cat: []

        diff = SnapshotDiff()  # no changes
        md = build_markdown(profile, [], today, author="A", role="R",
                            company="C", snapshot_diff=diff)
        assert "## Change History" not in md

    @patch("helpers.reporting.markdown.load_deadline_windows")
    def test_no_change_history_when_none(self, mock_cfg):
        mock_cfg.return_value = {
            "recent_completed_days": 7,
            "extended_completed_days": 30,
        }
        from helpers.reporting.markdown import build_markdown

        today = date(2026, 3, 25)
        profile = MagicMock()
        profile.tasks_for_category = lambda cat: []

        md = build_markdown(profile, [], today, author="A", role="R",
                            company="C", snapshot_diff=None)
        assert "## Change History" not in md


# ══════════════════════════════════════════════════════════════════════════════
#  Timelines Integrity Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestTimelinesIntegrity:
    """Test the timelines integrity checker."""

    def test_missing_timelines_sheet(self):
        from helpers.schema.integrity import check_timelines
        wb = MagicMock()
        wb.sheetnames = ["Projects", "Tasks"]
        report = check_timelines(wb)
        assert report.has_errors
        assert "missing" in report.issues[0].message.lower()

    def test_empty_timelines_sheet(self):
        from helpers.schema.integrity import check_timelines, IntegrityReport
        from openpyxl import Workbook

        wb = Workbook()
        wb.create_sheet("Timelines")
        # Also create source sheets (empty)
        wb.create_sheet("Projects")
        wb.create_sheet("Tasks")
        wb.create_sheet("Deliverables")

        report = check_timelines(wb)
        assert report.has_warnings or report.is_healthy

    def test_check_and_repair_calls_sync(self):
        from helpers.schema.integrity import check_and_repair

        wb = MagicMock()
        wb.sheetnames = ["Projects", "Tasks"]  # No Timelines

        with patch("helpers.schema.timelines.sync_timelines") as mock_sync, \
             patch("helpers.schema.gantt.build_gantt_sheet") as mock_gantt:
            # After sync, check again — now it has Timelines
            def add_timelines(wb_arg):
                wb.sheetnames = ["Projects", "Tasks", "Timelines"]
                return 5
            mock_sync.side_effect = add_timelines

            # Patch check_timelines to return error first, then OK
            with patch("helpers.schema.integrity.check_timelines") as mock_check:
                from helpers.schema.integrity import IntegrityReport, IntegrityIssue
                bad = IntegrityReport(issues=[
                    IntegrityIssue("error", None, None, "Timelines sheet is missing")
                ])
                good = IntegrityReport()
                mock_check.side_effect = [bad, good]

                report = check_and_repair(wb)
                assert report.repaired
                assert report.is_healthy


class TestIntegrityReport:
    """Test IntegrityReport data structure."""

    def test_healthy(self):
        from helpers.schema.integrity import IntegrityReport
        r = IntegrityReport()
        assert r.is_healthy
        assert not r.has_errors
        assert not r.has_warnings
        assert r.error_count == 0

    def test_with_errors(self):
        from helpers.schema.integrity import IntegrityReport, IntegrityIssue
        r = IntegrityReport(issues=[
            IntegrityIssue("error", 2, "Title", "Broken formula"),
            IntegrityIssue("warning", 3, "Status", "Empty cell"),
        ])
        assert r.has_errors
        assert r.has_warnings
        assert r.error_count == 1
        assert r.warning_count == 1
        assert not r.is_healthy
