"""Timelines sheet synchroniser.

Reads Projects / Tasks / Deliverables and writes corresponding rows to
the Timelines sheet.  Uses VLOOKUP formulas or direct cell references so
that updates to the source sheets propagate automatically.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from helpers.data.tasks import clean
from helpers.schema.sheets import (
    SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES, SHEET_TIMELINES,
    SHEET_META,
)
from helpers.schema.columns import (
    TIMELINES_COLUMNS, headers_for,
    column_index as ci,
)


# ── Timeline section configuration ────────────────────────────────────────────
# Each section maps source sheet fields to Timelines columns 3–14.
# A string entry means "VLOOKUP this field"; None means "leave blank".
# The special key "duration_formula" inserts an end−start formula.

_TIMELINE_SECTIONS = [
    {
        "label": "Project",
        "sheet": SHEET_PROJECTS,
        # Timelines cols: 3=Title, 4=Parent ID, 5=Start, 6=Duration,
        #   7=End, 8=Deadline, 9=Status, 10=%Complete, 11=TimeAlloc,
        #   12=TimeSpent, 13=SchedDate, 14=Milestones
        "columns": [
            ("Title",),             # col 3
            (None,),                # col 4 — projects have no parent
            ("Start Date",),        # col 5
            ("duration_formula",),  # col 6
            ("End Date",),          # col 7
            ("Deadline",),          # col 8
            ("Status",),            # col 9
            (None,),                # col 10
            (None,),                # col 11
            (None,),                # col 12
            (None,),                # col 13
            (None,),                # col 14
        ],
    },
    {
        "label": "Task",
        "sheet": SHEET_TASKS,
        "columns": [
            ("Title",),
            ("Project ID",),        # col 4 — parent
            ("Start Date",),
            ("duration_formula",),
            ("End Date",),
            ("Deadline",),
            ("Status",),
            (None,),
            (None,),
            (None,),
            ("Scheduled Date",),
            (None,),
        ],
    },
    {
        "label": "Deliverable",
        "sheet": SHEET_DELIVERABLES,
        "columns": [
            ("Title",),
            ("Task ID",),           # col 4 — parent
            ("Start Date",),
            ("duration_formula",),
            ("End Date",),
            ("Deadline",),
            ("Status",),
            ("% Complete",),
            ("Time Allocated",),
            ("Time Spent",),
            (None,),
            (None,),
        ],
    },
]


def _ensure_sheet(wb: Workbook, name: str):
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    return wb[name]


def _write_headers(ws):
    for i, col in enumerate(TIMELINES_COLUMNS, start=1):
        ws.cell(row=1, column=i, value=col.name)


def _clear_data(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)


def _vlookup(lookup_val_cell: str, table_range: str, col_idx: int) -> str:
    """Build a VLOOKUP formula string."""
    return f'=IFERROR(VLOOKUP({lookup_val_cell},{table_range},{col_idx},FALSE),"")'


def _table_range(wb: Workbook, sheet_name: str) -> str:
    """Return the VLOOKUP table range string for a sheet."""
    if sheet_name not in wb.sheetnames:
        return ""
    s = wb[sheet_name]
    max_col = get_column_letter(s.max_column)
    return f"'{sheet_name}'!$A:${max_col}"


def _write_section(ws, wb: Workbook, start_row: int, section: dict) -> int:
    """Write VLOOKUP rows for one entity type. Returns the next available row."""
    sheet_name = section["sheet"]
    if sheet_name not in wb.sheetnames:
        return start_row

    source_ws = wb[sheet_name]
    data_range = _table_range(wb, sheet_name)
    col_specs = section["columns"]

    # Pre-compute 1-based VLOOKUP column indices for each field
    field_indices: dict[str, int] = {}
    for spec in col_specs:
        field = spec[0]
        if field and field != "duration_formula" and field not in field_indices:
            field_indices[field] = ci(sheet_name, field) + 1

    row = start_row
    for r in range(2, source_ws.max_row + 1):
        item_id = clean(source_ws.cell(row=r, column=1).value)
        if not item_id:
            continue

        a_ref = f"$A{row}"
        ws.cell(row=row, column=1, value=item_id)
        ws.cell(row=row, column=2, value=section["label"])

        for col_offset, spec in enumerate(col_specs):
            col = col_offset + 3
            field = spec[0]
            if field is None:
                ws.cell(row=row, column=col, value="")
            elif field == "duration_formula":
                # Duration = End Date − Start Date (cols 7 and 5 in the timeline)
                ws.cell(row=row, column=col).value = (
                    f'=IFERROR(IF(F{row}="","",G{row}-F{row}),"")'
                )
            else:
                ws.cell(row=row, column=col).value = (
                    f'=IFERROR(VLOOKUP({a_ref},{data_range},'
                    f'{field_indices[field]},FALSE),"")'
                )

        row += 1

    return row


def sync_timelines(wb: Workbook) -> int:
    """Rebuild the Timelines sheet from the three data sheets.

    Returns the number of rows written (excluding header).

    Each Project / Task / Deliverable becomes a row.  Title, dates, and
    status columns are VLOOKUP formulas referencing the source sheets so
    edits in the data sheets propagate to Timelines automatically.
    """
    ws = _ensure_sheet(wb, SHEET_TIMELINES)
    _clear_data(ws)
    _write_headers(ws)

    # Apply sheet metadata
    meta = SHEET_META.get(SHEET_TIMELINES)
    if meta:
        if meta.frozen_pane:
            ws.freeze_panes = meta.frozen_pane
        if meta.tab_color:
            ws.sheet_properties.tabColor = meta.tab_color

    row = 2
    for section in _TIMELINE_SECTIONS:
        row = _write_section(ws, wb, row, section)

    return row - 2  # rows written
