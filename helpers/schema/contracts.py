"""Cross-sheet relationship contracts and validation utilities.

Defines the foreign-key relationships between data sheets and provides
functions to validate referential integrity across a workbook.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.workbook import Workbook

from .sheets import SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES, SHEET_TIMELINES
from .columns import column_index


# ── Relationship descriptors ──────────────────────────────────────────────────

@dataclass(frozen=True)
class Relationship:
    """Describes a one-to-many FK link between two sheets."""
    name: str
    parent_sheet: str
    parent_column: str      # column header in parent sheet
    child_sheet: str
    child_column: str       # FK column header in child sheet


RELATIONSHIPS: tuple[Relationship, ...] = (
    Relationship(
        name="ProjectToTasks",
        parent_sheet=SHEET_PROJECTS,
        parent_column="Project ID",
        child_sheet=SHEET_TASKS,
        child_column="Project ID",
    ),
    Relationship(
        name="TaskToDeliverables",
        parent_sheet=SHEET_TASKS,
        parent_column="Task ID",
        child_sheet=SHEET_DELIVERABLES,
        child_column="Task ID",
    ),
)


# ── Validation ─────────────────────────────────────────────────────────────────

def _collect_column_values(ws, col_idx: int) -> set[str]:
    """Collect all non-empty string values in *col_idx* (0-based), skipping header."""
    values: set[str] = set()
    for row in ws.iter_rows(min_row=2, min_col=col_idx + 1, max_col=col_idx + 1, values_only=True):
        v = row[0]
        if v is not None and str(v).strip():
            values.add(str(v).strip())
    return values


def validate_foreign_keys(wb: Workbook) -> list[str]:
    """Return a list of human-readable error strings for broken FK references.

    An empty list means all relationships are intact.
    """
    errors: list[str] = []
    for rel in RELATIONSHIPS:
        parent_ws = wb[rel.parent_sheet] if rel.parent_sheet in wb.sheetnames else None
        child_ws = wb[rel.child_sheet] if rel.child_sheet in wb.sheetnames else None

        if parent_ws is None or child_ws is None:
            continue  # sheets may not exist yet during migration

        parent_idx = column_index(rel.parent_sheet, rel.parent_column)
        child_idx = column_index(rel.child_sheet, rel.child_column)

        parent_ids = _collect_column_values(parent_ws, parent_idx)
        child_fks = _collect_column_values(child_ws, child_idx)

        orphans = child_fks - parent_ids
        for orphan in sorted(orphans):
            errors.append(
                f"[{rel.name}] {rel.child_sheet}.{rel.child_column} value "
                f"'{orphan}' has no matching row in {rel.parent_sheet}.{rel.parent_column}"
            )
    return errors


def validate_ids(wb: Workbook) -> list[str]:
    """Check for duplicate IDs in each data sheet's primary-key column.

    Returns a list of error strings (empty = valid).
    """
    from .ids import parse_id  # local import to avoid circular

    errors: list[str] = []
    pk_columns = {
        SHEET_PROJECTS:     "Project ID",
        SHEET_TASKS:        "Task ID",
        SHEET_DELIVERABLES: "Deliverable ID",
    }
    for sheet_name, pk_col in pk_columns.items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
        if ws is None:
            continue

        idx = column_index(sheet_name, pk_col)
        seen: dict[str, int] = {}
        for row_num, row in enumerate(
            ws.iter_rows(min_row=2, min_col=idx + 1, max_col=idx + 1, values_only=True),
            start=2,
        ):
            v = row[0]
            if v is None:
                continue
            v_str = str(v).strip()
            if not v_str:
                continue
            # Validate format
            try:
                parse_id(v_str)
            except ValueError:
                errors.append(f"[{sheet_name}] Row {row_num}: Invalid ID format '{v_str}'")
                continue
            if v_str in seen:
                errors.append(
                    f"[{sheet_name}] Duplicate ID '{v_str}' at rows {seen[v_str]} and {row_num}"
                )
            else:
                seen[v_str] = row_num
    return errors
