"""Create a blank template workbook with all 6 sheets properly set up.

Usage:
    python -m helpers.schema.template "path/to/new_workbook.xlsx"

Creates a workbook with:
  - Overview, Projects, Tasks, Deliverables, Timelines, Gantt Chart sheets
  - Headers, column widths, frozen panes, tab colors, data validation
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from helpers.schema.sheets import (
    ALL_SHEETS, SHEET_META, DATA_SHEETS,
    SHEET_OVERVIEW, SHEET_PROJECTS,
    ALL_CATEGORIES,
)
from helpers.schema.columns import (
    COLUMNS_BY_SHEET, Column, ColumnType,
    headers_for,
)


# ── Styles ─────────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="003DA5", end_color="003DA5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B3CDE3"),
    right=Side(style="thin", color="B3CDE3"),
    top=Side(style="thin", color="B3CDE3"),
    bottom=Side(style="thin", color="B3CDE3"),
)


def create_template(path: str | Path) -> Path:
    """Create a new workbook template with all schema sheets.

    Returns the output path.
    """
    path = Path(path)
    wb = Workbook()

    # Remove the default "Sheet" created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name in ALL_SHEETS:
        ws = wb.create_sheet(sheet_name)
        meta = SHEET_META.get(sheet_name)

        # Apply metadata
        if meta:
            if meta.frozen_pane:
                ws.freeze_panes = meta.frozen_pane
            if meta.tab_color:
                ws.sheet_properties.tabColor = meta.tab_color

        # Write headers and set column widths for sheets that have schemas
        columns = COLUMNS_BY_SHEET.get(sheet_name)
        if columns:
            for i, col_def in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=i, value=col_def.name)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
                ws.column_dimensions[get_column_letter(i)].width = col_def.width

            # Data validation for choice columns
            for i, col_def in enumerate(columns, start=1):
                if col_def.choices:
                    col_letter = get_column_letter(i)
                    dv = DataValidation(
                        type="list",
                        formula1='"' + ','.join(col_def.choices) + '"',
                        allow_blank=True,
                    )
                    dv.error = f"Please select from: {', '.join(col_def.choices)}"
                    dv.errorTitle = "Invalid Value"
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}1000")

    # Overview sheet — just a title placeholder
    overview_ws = wb[SHEET_OVERVIEW]
    overview_ws.merge_cells("B1:G1")
    cell = overview_ws.cell(row=1, column=2, value="Weekly Status Report")
    cell.font = Font(name="Calibri", bold=True, size=20, color="003DA5")
    overview_ws.column_dimensions["A"].width = 3
    overview_ws.column_dimensions["B"].width = 38

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python -m helpers.schema.template <output.xlsx>")
        sys.exit(1)
    out = create_template(sys.argv[1])
    print(f"Template workbook created: {out}")
