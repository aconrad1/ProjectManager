"""Gantt Chart sheet builder — dynamic Excel conditional formatting.

Generates columns F+ as date headers (daily or weekly), and applies
conditional formatting rules so cells fill automatically based on
each item's Start/End dates and Status.

Items with no start date are placed in a *No Scheduled Start* section
at the bottom so they remain visible in the workbook.
"""

from __future__ import annotations

from datetime import date, timedelta
from typing import Literal

from openpyxl.workbook import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from helpers.data.tasks import clean
from helpers.schema.sheets import (
    SHEET_GANTT, SHEET_TIMELINES, SHEET_META,
    SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES,
)
from helpers.schema.columns import GANTT_COLUMNS, column_index as ci


# ── Colour palette ─────────────────────────────────────────────────────────────

FILL_IN_PROGRESS = PatternFill(start_color="336BBF", end_color="336BBF", fill_type="solid")   # blue
FILL_COMPLETED   = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")   # green
FILL_OVERDUE     = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")   # red
FILL_NOT_STARTED = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")   # gray
FILL_HEADER      = PatternFill(start_color="003DA5", end_color="003DA5", fill_type="solid")
HEADER_FONT      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
DATE_FONT        = Font(name="Calibri", size=8, color="FFFFFF")
_THIN_BORDER     = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _ensure_sheet(wb: Workbook, name: str):
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    return wb[name]


def _clear(ws):
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row + 1)
    # Also clear existing conditional formatting
    ws.conditional_formatting._cf_rules.clear()


def _date_range(start: date, end: date, granularity: str) -> list[date]:
    """Generate a list of dates between *start* and *end*."""
    dates: list[date] = []
    cur = start
    step = timedelta(days=7) if granularity == "weekly" else timedelta(days=1)
    while cur <= end:
        dates.append(cur)
        cur += step
    return dates


def _classify_items(
    wb: Workbook,
    tl_ws,
) -> tuple[list[str], list[str]]:
    """Split Timelines item IDs into (scheduled, unscheduled).

    Since Timelines uses VLOOKUP formulas that openpyxl cannot evaluate,
    we read Start Date directly from the source data sheets (Projects,
    Tasks, Deliverables) to determine which items have a start date.

    Returns two lists of item-ID strings.
    """
    if tl_ws is None:
        return [], []

    # Collect all item IDs from the Timelines sheet
    all_ids: list[str] = []
    for r in range(2, tl_ws.max_row + 1):
        item_id = clean(tl_ws.cell(row=r, column=1).value)
        if item_id:
            all_ids.append(item_id)

    # Build a set of IDs that have a non-empty Start Date in source sheets
    has_start: set[str] = set()
    source_map = {
        SHEET_PROJECTS:     "Start Date",
        SHEET_TASKS:        "Start Date",
        SHEET_DELIVERABLES: "Start Date",
    }
    for sheet_name, col_name in source_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        src_ws = wb[sheet_name]
        col_idx = ci(sheet_name, col_name) + 1   # 1-based
        for r in range(2, src_ws.max_row + 1):
            item_id = clean(src_ws.cell(row=r, column=1).value)
            val = src_ws.cell(row=r, column=col_idx).value
            if item_id and val is not None and val != "":
                has_start.add(item_id)

    scheduled = [iid for iid in all_ids if iid in has_start]
    unscheduled = [iid for iid in all_ids if iid not in has_start]
    return scheduled, unscheduled


# ── Section header styling ─────────────────────────────────────────────────────

FILL_SECTION = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=9, color="333333")


def build_gantt_sheet(
    wb: Workbook,
    start_date: date | None = None,
    end_date: date | None = None,
    granularity: Literal["daily", "weekly"] = "weekly",
) -> int:
    """Build (or rebuild) the Gantt Chart sheet.

    Reads Item IDs from Timelines, creates date columns F+, and applies
    conditional formatting so cells fill based on Start/End/Status.

    Items with no start date are grouped under a *No Scheduled Start*
    section header at the bottom of the sheet.

    Returns the number of data rows written (excluding section headers).
    """
    ws = _ensure_sheet(wb, SHEET_GANTT)
    _clear(ws)

    # Apply tab metadata
    meta = SHEET_META.get(SHEET_GANTT)
    if meta and meta.tab_color:
        ws.sheet_properties.tabColor = meta.tab_color

    # Classify items by start-date presence
    tl_ws = wb[SHEET_TIMELINES] if SHEET_TIMELINES in wb.sheetnames else None
    scheduled, unscheduled = _classify_items(wb, tl_ws)

    # Determine date range
    if start_date is None:
        start_date = date.today() - timedelta(days=7)
    if end_date is None:
        end_date = start_date + timedelta(weeks=12)

    dates = _date_range(start_date, end_date, granularity)
    if not dates:
        return 0

    # ── Write fixed headers A-E ────────────────────────────────────────────
    fixed_headers = [c.name for c in GANTT_COLUMNS]
    for i, h in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set fixed column widths
    for i, col_def in enumerate(GANTT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_def.width

    # ── Write date headers F+ ──────────────────────────────────────────────
    date_start_col = len(fixed_headers) + 1  # column F = 6
    for j, d in enumerate(dates):
        col = date_start_col + j
        cell = ws.cell(row=1, column=col, value=d)
        cell.number_format = "MMM DD"
        cell.font = DATE_FONT
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        ws.column_dimensions[get_column_letter(col)].width = 4

    # Freeze panes: columns A-E and header row
    if meta and meta.frozen_pane:
        ws.freeze_panes = meta.frozen_pane

    # ── Helper: write a VLOOKUP data row ──────────────────────────────────
    tl_range = f"'{SHEET_TIMELINES}'!$A:$K"

    def _write_item_row(row: int, item_id: str) -> None:
        a_ref = f"$A{row}"
        ws.cell(row=row, column=1, value=item_id)
        ws.cell(row=row, column=2).value = f'=IFERROR(VLOOKUP({a_ref},{tl_range},3,FALSE),"")'
        ws.cell(row=row, column=3).value = f'=IFERROR(VLOOKUP({a_ref},{tl_range},5,FALSE),"")'
        ws.cell(row=row, column=4).value = f'=IFERROR(VLOOKUP({a_ref},{tl_range},7,FALSE),"")'
        ws.cell(row=row, column=5).value = f'=IFERROR(VLOOKUP({a_ref},{tl_range},9,FALSE),"")'

    # ── Write scheduled data rows ─────────────────────────────────────────
    for i, item_id in enumerate(scheduled):
        _write_item_row(i + 2, item_id)

    # ── "No Scheduled Start" section ──────────────────────────────────────
    next_row = len(scheduled) + 2
    if unscheduled:
        sec_cell = ws.cell(row=next_row, column=1, value="No Scheduled Start")
        sec_cell.font = SECTION_FONT
        sec_cell.fill = FILL_SECTION
        for col_i in range(2, len(fixed_headers) + 1):
            ws.cell(row=next_row, column=col_i).fill = FILL_SECTION
        next_row += 1

        for item_id in unscheduled:
            _write_item_row(next_row, item_id)
            next_row += 1

    total_items = len(scheduled) + len(unscheduled)

    # ── Conditional formatting for date cells ─────────────────────────────
    if total_items and dates:
        first_data_row = 2
        last_data_row = next_row - 1   # includes unscheduled rows
        first_date_col = get_column_letter(date_start_col)
        last_date_col = get_column_letter(date_start_col + len(dates) - 1)
        date_area = f"{first_date_col}{first_data_row}:{last_date_col}{last_data_row}"

        # Reference cells for first data row (anchored row with $)
        start_ref = f"$C{first_data_row}"   # Start date in col C
        end_ref = f"$D{first_data_row}"     # End date in col D
        status_ref = f"$E{first_data_row}"  # Status in col E
        date_cell_ref = f"{first_date_col}$1"  # Date header in row 1

        # Rule 1: Completed — green fill when date is between start and end AND status contains "completed"
        ws.conditional_formatting.add(
            date_area,
            FormulaRule(
                formula=[
                    f'AND({date_cell_ref}>={start_ref},{date_cell_ref}<={end_ref},'
                    f'SEARCH("complet",{status_ref})>0)'
                ],
                fill=FILL_COMPLETED,
            ),
        )

        # Rule 2: Overdue — red fill when date between start and end AND end < TODAY AND status not completed
        ws.conditional_formatting.add(
            date_area,
            FormulaRule(
                formula=[
                    f'AND({date_cell_ref}>={start_ref},{date_cell_ref}<={end_ref},'
                    f'{end_ref}<TODAY(),ISERROR(SEARCH("complet",{status_ref})))'
                ],
                fill=FILL_OVERDUE,
            ),
        )

        # Rule 3: In Progress — blue fill when date is between start and end
        ws.conditional_formatting.add(
            date_area,
            FormulaRule(
                formula=[
                    f'AND({date_cell_ref}>={start_ref},{date_cell_ref}<={end_ref},'
                    f'ISERROR(SEARCH("complet",{status_ref})))'
                ],
                fill=FILL_IN_PROGRESS,
            ),
        )

    return total_items
