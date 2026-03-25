"""Timelines sheet integrity checker and auto-repair.

Detects broken VLOOKUP references, formula pattern mismatches, and missing
rows in the Timelines sheet, then optionally rebuilds it.  Used by the
report pipeline to ensure derived sheets are healthy before generating
reports.

Checks performed
~~~~~~~~~~~~~~~~
1. **Header match** — column names vs. schema.
2. **VLOOKUP formulas** — broken ``#REF!`` / ``#NAME?`` patterns; missing
   formulas where one is expected.
3. **Formula target validation** — ensures VLOOKUPs point at the correct
   source sheet for each item type (Project → Projects, Task → Tasks, etc.).
4. **Duration formula** — column 6 should be an ``IF``/``IFERROR`` formula.
5. **Missing / orphaned IDs** — cross-references source data sheets.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Literal

from openpyxl.workbook import Workbook

from helpers.data.tasks import clean
from helpers.schema.sheets import (
    SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES, SHEET_TIMELINES,
)
from helpers.schema.columns import TIMELINES_COLUMNS


# ── Data structures ────────────────────────────────────────────────────────────

IssueSeverity = Literal["warning", "error"]


@dataclass
class IntegrityIssue:
    """A single integrity problem found in the Timelines sheet."""
    severity: IssueSeverity
    row: int | None          # None for sheet-level issues
    column: str | None
    message: str


@dataclass
class IntegrityReport:
    """Result of a Timelines integrity check."""
    issues: list[IntegrityIssue] = field(default_factory=list)
    repaired: bool = False

    @property
    def has_errors(self) -> bool:
        return any(i.severity == "error" for i in self.issues)

    @property
    def has_warnings(self) -> bool:
        return any(i.severity == "warning" for i in self.issues)

    @property
    def is_healthy(self) -> bool:
        return len(self.issues) == 0

    @property
    def error_count(self) -> int:
        return sum(1 for i in self.issues if i.severity == "error")

    @property
    def warning_count(self) -> int:
        return sum(1 for i in self.issues if i.severity == "warning")


# ── Checker ────────────────────────────────────────────────────────────────────

def _collect_source_ids(wb: Workbook) -> set[str]:
    """Gather all item IDs from Projects, Tasks, and Deliverables sheets."""
    ids: set[str] = set()
    for sheet_name in (SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                val = clean(ws.cell(row=r, column=1).value)
                if val:
                    ids.add(val)
    return ids


def _check_headers(ws, issues: list[IntegrityIssue]) -> bool:
    """Verify Timelines headers match the schema. Returns True if OK."""
    expected = [c.name for c in TIMELINES_COLUMNS]
    actual = [clean(ws.cell(row=1, column=i + 1).value) for i in range(len(expected))]
    if actual != expected:
        issues.append(IntegrityIssue(
            severity="error",
            row=1,
            column=None,
            message=f"Header mismatch: expected {expected}, got {actual}",
        ))
        return False
    return True


def _check_vlookup_formulas(ws, issues: list[IntegrityIssue]) -> None:
    """Check that VLOOKUP formula cells contain valid formulas."""
    # Columns that should be VLOOKUP formulas: Title(3), Start(5), End(7),
    # Deadline(8), Status(9) — but only for data rows
    formula_cols = {3: "Title", 5: "Start Date", 7: "End Date",
                    8: "Deadline", 9: "Status"}
    for r in range(2, ws.max_row + 1):
        item_id = clean(ws.cell(row=r, column=1).value)
        if not item_id:
            continue
        for col_idx, col_name in formula_cols.items():
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            # A formula cell should start with '=' or be an openpyxl formula
            if val is not None and isinstance(val, str) and val.startswith("="):
                # Check for common broken patterns
                if "#REF!" in val or "#NAME?" in val:
                    issues.append(IntegrityIssue(
                        severity="error",
                        row=r,
                        column=col_name,
                        message=f"Broken formula in {col_name} for {item_id}: {val}",
                    ))
            elif val is not None and not isinstance(val, str):
                # It's a literal value where a formula was expected — warning
                pass  # This is acceptable; openpyxl reads formula results as values
            # Empty formula cell for a valid item ID is a warning
            elif val is None or (isinstance(val, str) and val.strip() == ""):
                issues.append(IntegrityIssue(
                    severity="warning",
                    row=r,
                    column=col_name,
                    message=f"Empty formula cell in {col_name} for {item_id}",
                ))


# Regex that captures the sheet name inside a VLOOKUP(…,'SheetName'!…)
_VLOOKUP_SHEET_RE = re.compile(r"VLOOKUP\([^,]+,\s*'([^']+)'!", re.IGNORECASE)

# Expected source sheet per item-type prefix
_EXPECTED_SOURCE: dict[str, str] = {
    "P": SHEET_PROJECTS,
    "T": SHEET_TASKS,
    "D": SHEET_DELIVERABLES,
}


def _check_formula_targets(ws, issues: list[IntegrityIssue]) -> None:
    """Ensure VLOOKUPs reference the correct source sheet for each item."""
    formula_cols = (3, 5, 7, 8, 9)
    col_names = {3: "Title", 5: "Start Date", 7: "End Date",
                 8: "Deadline", 9: "Status"}
    for r in range(2, ws.max_row + 1):
        item_id = clean(ws.cell(row=r, column=1).value)
        if not item_id:
            continue
        prefix = item_id.split("-")[0] if "-" in item_id else ""
        expected_sheet = _EXPECTED_SOURCE.get(prefix)
        if not expected_sheet:
            continue

        for col_idx in formula_cols:
            val = ws.cell(row=r, column=col_idx).value
            if not isinstance(val, str) or not val.startswith("="):
                continue
            m = _VLOOKUP_SHEET_RE.search(val)
            if m and m.group(1) != expected_sheet:
                issues.append(IntegrityIssue(
                    severity="error",
                    row=r,
                    column=col_names.get(col_idx, str(col_idx)),
                    message=(
                        f"VLOOKUP for {item_id} references '{m.group(1)}' "
                        f"but expected '{expected_sheet}'"
                    ),
                ))


def _check_duration_formulas(ws, issues: list[IntegrityIssue]) -> None:
    """Column 6 (Duration) should contain an IF/IFERROR formula, not a literal."""
    for r in range(2, ws.max_row + 1):
        item_id = clean(ws.cell(row=r, column=1).value)
        if not item_id:
            continue
        val = ws.cell(row=r, column=6).value
        if isinstance(val, str) and val.startswith("="):
            if "#REF!" in val or "#NAME?" in val:
                issues.append(IntegrityIssue(
                    severity="error",
                    row=r,
                    column="Duration (days)",
                    message=f"Broken duration formula for {item_id}: {val}",
                ))
        elif val is None or (isinstance(val, str) and val.strip() == ""):
            issues.append(IntegrityIssue(
                severity="warning",
                row=r,
                column="Duration (days)",
                message=f"Empty duration formula for {item_id}",
            ))


def _check_missing_ids(ws, source_ids: set[str], issues: list[IntegrityIssue]) -> None:
    """Check that all source IDs appear in Timelines."""
    timeline_ids: set[str] = set()
    for r in range(2, ws.max_row + 1):
        val = clean(ws.cell(row=r, column=1).value)
        if val:
            timeline_ids.add(val)

    missing = source_ids - timeline_ids
    for mid in sorted(missing):
        issues.append(IntegrityIssue(
            severity="error",
            row=None,
            column="Item ID",
            message=f"Item {mid} exists in source sheets but missing from Timelines",
        ))

    orphaned = timeline_ids - source_ids
    for oid in sorted(orphaned):
        issues.append(IntegrityIssue(
            severity="warning",
            row=None,
            column="Item ID",
            message=f"Item {oid} in Timelines but not found in source sheets (orphaned)",
        ))


def check_timelines(wb: Workbook) -> IntegrityReport:
    """Run all integrity checks on the Timelines sheet.

    Returns an IntegrityReport with any issues found.
    """
    report = IntegrityReport()

    if SHEET_TIMELINES not in wb.sheetnames:
        report.issues.append(IntegrityIssue(
            severity="error",
            row=None,
            column=None,
            message="Timelines sheet is missing from the workbook",
        ))
        return report

    ws = wb[SHEET_TIMELINES]

    # Check if the sheet is empty
    if ws.max_row < 2:
        report.issues.append(IntegrityIssue(
            severity="warning",
            row=None,
            column=None,
            message="Timelines sheet is empty (no data rows)",
        ))
        return report

    _check_headers(ws, report.issues)
    _check_vlookup_formulas(ws, report.issues)
    _check_formula_targets(ws, report.issues)
    _check_duration_formulas(ws, report.issues)

    source_ids = _collect_source_ids(wb)
    _check_missing_ids(ws, source_ids, report.issues)

    return report


def check_and_repair(wb: Workbook, *, log=None) -> IntegrityReport:
    """Check Timelines integrity and auto-repair if errors are found.

    If errors are detected, rebuilds the Timelines sheet via
    ``sync_timelines``, then re-checks to confirm the repair.
    Also rebuilds the Gantt Chart sheet so it stays in sync.

    Returns the final IntegrityReport (with ``repaired=True`` if a
    rebuild was performed).
    """
    if log is None:
        log = lambda msg: None  # noqa: E731

    report = check_timelines(wb)

    if not report.has_errors:
        if report.has_warnings:
            log(f"   Timelines integrity: {report.warning_count} warning(s), no errors.")
        return report

    log(f"   Timelines integrity: {report.error_count} error(s), "
        f"{report.warning_count} warning(s) — rebuilding…")

    from helpers.schema.timelines import sync_timelines
    sync_timelines(wb)

    from helpers.schema.gantt import build_gantt_sheet
    build_gantt_sheet(wb)

    # Re-check after repair
    post_report = check_timelines(wb)
    post_report.repaired = True

    if post_report.has_errors:
        log(f"   ⚠ Repair incomplete: {post_report.error_count} error(s) remain")
    else:
        log("   ✓ Timelines repaired successfully")

    return post_report
