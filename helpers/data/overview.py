"""Overview tab writer — populates the Excel 'Overview' sheet.

Builds a compact dashboard with five sections:
1. Project Status Summary — one row per active project
2. Weekly Schedule Grid  — 7-day × 5-priority matrix of scheduled tasks
3. Time Overview         — allocated / spent / remaining per project
4. Upcoming Deadlines    — configurable window (from deadlines.json)
5. Change History        — diff against previous snapshot (if provided)
"""

from __future__ import annotations

from datetime import date, timedelta
from collections import Counter

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from helpers.config.loader import (
    load_deadline_windows,
    priority_labels as _load_priority_labels,
    status_bg_color,
    active_categories as _load_active_categories,
    terminal_statuses as _load_terminal_statuses,
)
from helpers.domain.profile import Profile
from helpers.domain.task import Task
from helpers.reporting.snapshot_diff import SnapshotDiff
from helpers.schema.sheets import SHEET_OVERVIEW

# ── Brand colours (hex, no '#') ───────────────────────────────────────────────
BRAND_BLUE_DARK = "003DA5"
BRAND_BLUE_MID = "336BBF"
BRAND_BLUE_LIGHT = "B3CDE3"
BRAND_BLUE_WASH = "E6EFF8"
BRAND_WHITE = "FFFFFF"

# ── Reusable styles ───────────────────────────────────────────────────────────
_THIN_BORDER = Border(
    left=Side(style="thin", color=BRAND_BLUE_LIGHT),
    right=Side(style="thin", color=BRAND_BLUE_LIGHT),
    top=Side(style="thin", color=BRAND_BLUE_LIGHT),
    bottom=Side(style="thin", color=BRAND_BLUE_LIGHT),
)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=BRAND_WHITE)
HEADER_FILL = PatternFill(start_color=BRAND_BLUE_DARK, end_color=BRAND_BLUE_DARK, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=BRAND_BLUE_WASH, end_color=BRAND_BLUE_WASH, fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=14, color=BRAND_BLUE_DARK)
SUBSECTION_FONT = Font(name="Calibri", bold=True, size=12, color=BRAND_BLUE_MID)
BODY_FONT = Font(name="Calibri", size=10)
BODY_WRAP = Alignment(wrap_text=True, vertical="top")
CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_FONT = Font(name="Calibri", bold=True, size=20, color=BRAND_BLUE_DARK)
SUBTITLE_FONT = Font(name="Calibri", size=12, color=BRAND_BLUE_MID)

# Schedule-grid status colours — built from dimension table at import time
def _build_status_fills():
    from helpers.config.loader import load
    fills = {}
    for s in load("status")["values"]:
        bg = s["bg_color"].lstrip("#")
        fills[s["name"].lower()] = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    return fills

_STATUS_FILLS = _build_status_fills()

PRIORITY_LABELS = _load_priority_labels()


def _clear_sheet(ws):
    for merge in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(merge))
        except KeyError:
            ws.merged_cells.ranges.discard(merge)
    ws.delete_rows(1, ws.max_row + 1)


def _write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTRE
        cell.border = _THIN_BORDER


def _write_data_row(ws, row, values, col_start=1, alt=False):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        cell.font = BODY_FONT
        cell.alignment = BODY_WRAP
        cell.border = _THIN_BORDER
        if alt:
            cell.fill = ALT_ROW_FILL


def _set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _fmt_hours(val: float | None) -> str:
    if val is None or val == 0:
        return "—"
    return f"{val:.1f}"


def write_overview(
    wb: Workbook,
    profile: Profile,
    moved_titles: list[str],
    today: date | None = None,
    *,
    author: str = "",
    role: str = "",
    company: str = "",
    snapshot_diff: SnapshotDiff | None = None,
    deadline_windows: dict[str, int] | None = None,
) -> None:
    """Populate the Overview tab with a compact dashboard."""
    if today is None:
        today = date.today()

    # Derive data from domain hierarchy
    weekly_tasks = profile.tasks_for_category("Weekly")
    ongoing_tasks = profile.tasks_for_category("Ongoing")
    _active_cats = {c.lower() for c in _load_active_categories()}
    active_projects = [
        p for p in profile.projects
        if p.category.strip().lower() in _active_cats
    ]

    ws = wb[SHEET_OVERVIEW]
    _clear_sheet(ws)
    _set_col_widths(ws, {1: 3, 2: 36, 3: 14, 4: 14, 5: 12, 6: 14, 7: 14, 8: 14, 9: 14})

    row = 1

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws.cell(row=row, column=2, value=f"Weekly Dashboard — {today.strftime('%B %d, %Y')}").font = TITLE_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws.cell(row=row, column=2, value=f"{author}  |  {role}  |  {company}").font = SUBTITLE_FONT
    row += 2

    # ══════════════════════════════════════════════════════════════════════
    #  SECTION 1 — Project Status Summary
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws.cell(row=row, column=2, value="Project Status Summary").font = SECTION_FONT
    row += 1

    headers = ["Project", "Category", "Status", "Priority", "Active Tasks",
               "% Complete", "Allocated (hrs)", "Spent (hrs)"]
    _write_header_row(ws, row, headers, col_start=2)
    row += 1

    for idx, proj in enumerate(sorted(active_projects, key=lambda p: min((t.priority for t in p.tasks), default=5))):
        _terminal_sts = {s.lower() for s in _load_terminal_statuses()} | {"on hold"}
        active_count = sum(
            1 for t in proj.tasks
            if t.status.strip().lower() not in _terminal_sts
        )
        total_deliverables = sum(len(t.deliverables) for t in proj.tasks)
        done_deliverables = sum(
            1 for t in proj.tasks for d in t.deliverables
            if d.status.strip().lower() in {s.lower() for s in _load_terminal_statuses()}
        )
        pct = f"{done_deliverables}/{total_deliverables}" if total_deliverables else "—"
        alloc = proj.time_allocated_total
        spent = proj.time_spent_total
        top_pri = min((t.priority for t in proj.tasks), default=5)
        _write_data_row(ws, row, [
            proj.title, proj.category, proj.status,
            PRIORITY_LABELS.get(top_pri, f"P{top_pri}"),
            active_count, pct, _fmt_hours(alloc), _fmt_hours(spent),
        ], col_start=2, alt=idx % 2 == 1)
        row += 1
    row += 1

    # ══════════════════════════════════════════════════════════════════════
    #  SECTION 2 — Weekly Schedule Grid (7 days × 5 priority levels)
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws.cell(row=row, column=2, value="Weekly Schedule").font = SECTION_FONT
    row += 1

    # Build a lookup: (day_offset, priority) → [task, ...]
    schedule_lookup: dict[tuple[int, int], list[Task]] = {}
    _excluded = {s.lower() for s in _load_terminal_statuses()} | {"on hold"}
    all_active = [t for t in weekly_tasks + ongoing_tasks
                  if t.status.strip().lower() not in _excluded]
    for t in all_active:
        if t.scheduled_date is not None:
            offset = (t.scheduled_date - today).days
            if 0 <= offset < 7:
                schedule_lookup.setdefault((offset, t.priority), []).append(t)

    # Day header row
    day_headers = ["Priority"]
    for d in range(7):
        day = today + timedelta(days=d)
        day_headers.append(day.strftime("%a %m/%d"))
    _write_header_row(ws, row, day_headers, col_start=2)
    row += 1

    # One row per priority level
    for pri in range(1, 6):
        label = PRIORITY_LABELS.get(pri, f"P{pri}")
        ws.cell(row=row, column=2, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2).border = _THIN_BORDER
        for d in range(7):
            col = 3 + d
            tasks_in_slot = schedule_lookup.get((d, pri), [])
            if tasks_in_slot:
                cell_text = "\n".join(t.title for t in tasks_in_slot)
                cell = ws.cell(row=row, column=col, value=cell_text)
                cell.font = BODY_FONT
                cell.alignment = CENTRE
                cell.border = _THIN_BORDER
                # Use the first task's status for cell colour
                status_key = tasks_in_slot[0].status.strip().lower()
                fill = _STATUS_FILLS.get(status_key)
                if fill:
                    cell.fill = fill
            else:
                cell = ws.cell(row=row, column=col, value="")
                cell.font = BODY_FONT
                cell.alignment = CENTRE
                cell.border = _THIN_BORDER
                if (d + pri) % 2 == 0:
                    cell.fill = ALT_ROW_FILL
        row += 1
    row += 1

    # ══════════════════════════════════════════════════════════════════════
    #  SECTION 3 — Time Overview
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws.cell(row=row, column=2, value="Time Overview").font = SECTION_FONT
    row += 1

    time_headers = ["Project", "Allocated (hrs)", "Spent (hrs)",
                    "Remaining (hrs)", "Utilization %"]
    _write_header_row(ws, row, time_headers, col_start=2)
    row += 1

    total_alloc = 0.0
    total_spent = 0.0
    for idx, proj in enumerate(active_projects):
        alloc = proj.time_allocated_total
        spent = proj.time_spent_total
        remaining = max(0.0, alloc - spent) if alloc else 0.0
        util = f"{spent / alloc * 100:.0f}%" if alloc else "—"
        total_alloc += alloc
        total_spent += spent
        _write_data_row(ws, row, [
            proj.title, _fmt_hours(alloc), _fmt_hours(spent),
            _fmt_hours(remaining), util,
        ], col_start=2, alt=idx % 2 == 1)
        row += 1

    # Totals row
    total_remaining = max(0.0, total_alloc - total_spent)
    total_util = f"{total_spent / total_alloc * 100:.0f}%" if total_alloc else "—"
    for i, v in enumerate(["TOTAL", _fmt_hours(total_alloc), _fmt_hours(total_spent),
                           _fmt_hours(total_remaining), total_util]):
        cell = ws.cell(row=row, column=2 + i, value=v)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill(start_color=BRAND_BLUE_LIGHT, end_color=BRAND_BLUE_LIGHT, fill_type="solid")
        cell.border = _THIN_BORDER
    row += 2

    # ══════════════════════════════════════════════════════════════════════
    #  SECTION 4 — Upcoming Deadlines (configurable window)
    # ══════════════════════════════════════════════════════════════════════
    dl_cfg = deadline_windows or load_deadline_windows()
    deadline_days = dl_cfg.get("upcoming_deadline_days", 14)

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws.cell(row=row, column=2, value=f"Upcoming Deadlines (Next {deadline_days} Days)").font = SECTION_FONT
    row += 1

    cutoff = today + timedelta(days=deadline_days)
    upcoming: list[tuple[date, str, str, str]] = []
    for proj in active_projects:
        for task in proj.tasks:
            if task.deadline and today <= task.deadline <= cutoff:
                upcoming.append((task.deadline, task.title, task.status, proj.title))
            for d in task.deliverables:
                if d.deadline and today <= d.deadline <= cutoff:
                    upcoming.append((d.deadline, d.title, d.status, task.title))
    upcoming.sort(key=lambda x: x[0])

    if upcoming:
        _write_header_row(ws, row, ["Deadline", "Item", "Status", "Parent"], col_start=2)
        row += 1
        for idx, (dl, title, status, parent) in enumerate(upcoming):
            _write_data_row(ws, row, [
                dl.strftime("%Y-%m-%d"), title, status, parent,
            ], col_start=2, alt=idx % 2 == 1)
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=2,
                       value=f"No deadlines within the next {deadline_days} days.")
        cell.font = Font(name="Calibri", size=10, italic=True, color="555555")
        cell.alignment = BODY_WRAP
        row += 1
    row += 1

    # ══════════════════════════════════════════════════════════════════════
    #  SECTION 5 — Change History (snapshot diff)
    # ══════════════════════════════════════════════════════════════════════
    if snapshot_diff is not None and snapshot_diff.has_changes:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        ws.cell(row=row, column=2, value="Change History").font = SECTION_FONT
        row += 1

        _write_header_row(ws, row, ["Change", "Type", "ID", "Title", "Details"], col_start=2)
        row += 1

        for idx, change in enumerate(snapshot_diff.changes):
            detail = ""
            if change.kind == "modified":
                detail = "; ".join(f"{fc.field}: {fc.old} → {fc.new}" for fc in change.fields)
            _write_data_row(ws, row, [
                change.kind.title(),
                change.entity_type.title(),
                change.entity_id,
                change.title,
                detail[:120] if detail else "—",
            ], col_start=2, alt=idx % 2 == 1)
            row += 1
