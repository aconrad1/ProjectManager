"""Workbook I/O — load, read, and save helpers.

Wraps openpyxl so all other modules get data via pure function calls.
Uses the schema module as the single source of truth for sheet names
and column layouts.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook

from helpers.data.tasks import clean, parse_priority, parse_date, parse_percent
from helpers.schema.sheets import (
    SHEET_OVERVIEW,
    SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES,
    SHEET_TIMELINES, SHEET_GANTT,
)
from helpers.schema.columns import (
    PROJECTS_COLUMNS, TASKS_COLUMNS, DELIVERABLES_COLUMNS,
    headers_for, column_index,
)


# ── Schema-based reading ──────────────────────────────────────────────────────

def _read_sheet_rows(ws, columns) -> list[dict]:
    """Read all data rows from a sheet using schema-defined columns.

    Returns a list of dicts keyed by column name.
    """
    rows: list[dict] = []
    col_count = len(columns)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        padded = list(row) + [None] * (col_count - len(row))
        record = {}
        for i, col_def in enumerate(columns):
            record[col_def.name] = padded[i]
        rows.append(record)
    return rows


def load_projects(wb: Workbook) -> list[dict]:
    """Read all rows from the Projects sheet as dicts."""
    if SHEET_PROJECTS not in wb.sheetnames:
        return []
    return _read_sheet_rows(wb[SHEET_PROJECTS], PROJECTS_COLUMNS)


def load_tasks(wb: Workbook) -> list[dict]:
    """Read all rows from the Tasks sheet as dicts."""
    if SHEET_TASKS not in wb.sheetnames:
        return []
    return _read_sheet_rows(wb[SHEET_TASKS], TASKS_COLUMNS)


def load_deliverables(wb: Workbook) -> list[dict]:
    """Read all rows from the Deliverables sheet as dicts."""
    if SHEET_DELIVERABLES not in wb.sheetnames:
        return []
    return _read_sheet_rows(wb[SHEET_DELIVERABLES], DELIVERABLES_COLUMNS)


def load_all_data(wb: Workbook) -> tuple[list[dict], list[dict], list[dict]]:
    """Return ``(projects, tasks, deliverables)`` as lists of row dicts."""
    return load_projects(wb), load_tasks(wb), load_deliverables(wb)


# ── Workbook open / save ──────────────────────────────────────────────────────

def load_workbook(path: str | Path) -> Workbook:
    """Load the master workbook (keeping styles intact)."""
    return openpyxl.load_workbook(str(path))


def save_workbook(wb: Workbook, path: str | Path) -> None:
    """Save the workbook in-place."""
    wb.save(str(path))


def save_snapshot(wb: Workbook, workbook_path: str | Path, dest: Path) -> Path:
    """Save workbook in-place and copy a dated snapshot to *dest*.

    Returns the snapshot path.
    """
    wb.save(str(workbook_path))
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(workbook_path), str(dest))
    return dest
