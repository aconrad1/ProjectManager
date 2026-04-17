"""Completion detection — mark completed tasks/projects in-place.

Completed items stay on their sheet — the Projects sheet's Category
column is updated and a Date Completed is stamped.
"""

from __future__ import annotations

from datetime import date

from openpyxl.workbook import Workbook

from helpers.data.tasks import clean
from helpers.schema.sheets import SHEET_PROJECTS, SHEET_TASKS
from helpers.schema.columns import column_index
from helpers.config.loader import completion_aliases, terminal_statuses, terminal_categories


def _is_completed(status: str) -> bool:
    return status.lower().strip() in completion_aliases()


# ── New-schema completion (in-place update) ────────────────────────────────────

def process_completions(wb: Workbook, today: date | None = None) -> list[str]:
    """Scan Tasks sheet for completed items — stamp Date Completed in-place.

    Also updates the parent Project's Category to 'Completed' if all its
    tasks are completed.  Returns titles of newly completed items.
    """
    if today is None:
        today = date.today()

    moved: list[str] = []

    # --- Tasks sheet ---
    if SHEET_TASKS in wb.sheetnames:
        ws = wb[SHEET_TASKS]
        status_col = column_index(SHEET_TASKS, "Status") + 1      # 1-based
        title_col = column_index(SHEET_TASKS, "Title") + 1
        date_col = column_index(SHEET_TASKS, "Date Completed") + 1

        for row_idx in range(2, ws.max_row + 1):
            status_val = clean(ws.cell(row=row_idx, column=status_col).value)
            if _is_completed(status_val):
                existing = ws.cell(row=row_idx, column=date_col).value
                if existing is None:
                    ws.cell(row=row_idx, column=date_col, value=today)
                    title = clean(ws.cell(row=row_idx, column=title_col).value)
                    if title:
                        moved.append(title)

    # --- Projects sheet (auto-complete if all tasks done) ---
    if SHEET_PROJECTS in wb.sheetnames and SHEET_TASKS in wb.sheetnames:
        proj_ws = wb[SHEET_PROJECTS]
        task_ws = wb[SHEET_TASKS]

        proj_id_col = column_index(SHEET_PROJECTS, "Project ID") + 1
        proj_status_col = column_index(SHEET_PROJECTS, "Status") + 1
        proj_cat_col = column_index(SHEET_PROJECTS, "Category") + 1
        proj_date_col = column_index(SHEET_PROJECTS, "Date Completed") + 1

        task_proj_id_col = column_index(SHEET_TASKS, "Project ID") + 1
        task_status_col = column_index(SHEET_TASKS, "Status") + 1

        # Build a map: project_id → list of task statuses
        proj_tasks: dict[str, list[str]] = {}
        for row_idx in range(2, task_ws.max_row + 1):
            pid = clean(task_ws.cell(row=row_idx, column=task_proj_id_col).value)
            st = clean(task_ws.cell(row=row_idx, column=task_status_col).value)
            if pid:
                proj_tasks.setdefault(pid, []).append(st)

        for row_idx in range(2, proj_ws.max_row + 1):
            pid = clean(proj_ws.cell(row=row_idx, column=proj_id_col).value)
            if not pid:
                continue
            task_statuses = proj_tasks.get(pid, [])
            if task_statuses and all(_is_completed(s) for s in task_statuses):
                existing_date = proj_ws.cell(row=row_idx, column=proj_date_col).value
                if existing_date is None:
                    _completed_status = next(iter(terminal_statuses()))
                    _completed_cat = next(iter(terminal_categories()))
                    proj_ws.cell(row=row_idx, column=proj_status_col, value=_completed_status)
                    proj_ws.cell(row=row_idx, column=proj_cat_col, value=_completed_cat)
                    proj_ws.cell(row=row_idx, column=proj_date_col, value=today)

    return moved
