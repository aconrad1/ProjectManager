"""Thin abstraction for reading/writing worksheet rows by field name."""

from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from helpers.schema.columns import column_index


def _clean(value: Any) -> str:
    """Normalize a cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


class SheetAccessor:
    """Read and write worksheet rows by named fields."""

    def __init__(self, ws: Worksheet, sheet_name: str) -> None:
        self.ws = ws
        self._sheet_name = sheet_name
        self._col_cache: dict[str, int] = {}

    def _col(self, field: str) -> int:
        if field not in self._col_cache:
            self._col_cache[field] = column_index(self._sheet_name, field) + 1
        return self._col_cache[field]

    def get(self, row: int, field: str) -> str:
        return _clean(self.ws.cell(row=row, column=self._col(field)).value)

    def get_raw(self, row: int, field: str) -> Any:
        return self.ws.cell(row=row, column=self._col(field)).value

    def set(self, row: int, field: str, value: Any) -> None:
        self.ws.cell(row=row, column=self._col(field), value=value)

    def get_id(self, row: int) -> str:
        return _clean(self.ws.cell(row=row, column=1).value)

    def find_row(self, target_id: str) -> int | None:
        for row in range(2, self.ws.max_row + 1):
            if self.get_id(row) == target_id:
                return row
        return None

    def iter_rows(self):
        for row in range(2, self.ws.max_row + 1):
            if self.get_id(row):
                yield row
