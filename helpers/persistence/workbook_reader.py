"""Read an openpyxl Workbook into the domain hierarchy.

Reads the 6-sheet schema (Projects / Tasks / Deliverables) into the
Profile → Project → Task → Deliverable domain tree.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from helpers.domain.profile import Profile
from helpers.domain.project import Project
from helpers.domain.task import Task
from helpers.domain.deliverable import Deliverable
from helpers.data.tasks import clean, parse_priority, parse_date, parse_percent, parse_float
from helpers.schema.sheets import (
    SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES,
)
from helpers.schema.columns import (
    PROJECTS_COLUMNS, TASKS_COLUMNS, DELIVERABLES_COLUMNS,
    column_index,
)


# ── Helpers ────────────────────────────────────────────────────────────────────


def _row_dict(row, columns):
    """Map a values_only row tuple to a dict keyed by column name."""
    padded = list(row) + [None] * (len(columns) - len(row))
    return {col.name: padded[i] for i, col in enumerate(columns)}


# ── New-schema reader ──────────────────────────────────────────────────────────

def _read_projects(wb: Workbook) -> list[Project]:
    ws = wb[SHEET_PROJECTS]
    projects: list[Project] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        d = _row_dict(row, PROJECTS_COLUMNS)
        pid = clean(d.get("Project ID"))
        if not pid:
            continue
        projects.append(Project(
            id=pid,
            title=clean(d.get("Title")),
            project_id=pid,
            category=clean(d.get("Category")),
            description=clean(d.get("Description")),
            status=clean(d.get("Status")),
            supervisor=clean(d.get("Supervisor")),
            site=clean(d.get("Site")),
            priority=parse_priority(d.get("Priority")),
            notes=clean(d.get("Notes")),
            start=parse_date(d.get("Start Date")),
            end=parse_date(d.get("End Date")),
            deadline=parse_date(d.get("Deadline")),
            date_completed=parse_date(d.get("Date Completed")),
        ))
    return projects


def _read_domain_tasks(wb: Workbook) -> list[Task]:
    ws = wb[SHEET_TASKS]
    tasks: list[Task] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        d = _row_dict(row, TASKS_COLUMNS)
        tid = clean(d.get("Task ID"))
        if not tid:
            continue
        tasks.append(Task(
            id=tid,
            title=clean(d.get("Title")),
            task_id=tid,
            project_id=clean(d.get("Project ID")),
            description=clean(d.get("Description")),
            supervisor=clean(d.get("Supervisor")),
            site=clean(d.get("Site")),
            status=clean(d.get("Status")),
            priority=parse_priority(d.get("Priority")),
            start=parse_date(d.get("Start Date")),
            end=parse_date(d.get("End Date")),
            deadline=parse_date(d.get("Deadline")),
            commentary=clean(d.get("Status Commentary")),
            date_completed=parse_date(d.get("Date Completed")),
            scheduled_date=parse_date(d.get("Scheduled Date")),
        ))
    return tasks


def _read_deliverables(wb: Workbook) -> list[Deliverable]:
    if SHEET_DELIVERABLES not in wb.sheetnames:
        return []
    ws = wb[SHEET_DELIVERABLES]
    deliverables: list[Deliverable] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        d = _row_dict(row, DELIVERABLES_COLUMNS)
        did = clean(d.get("Deliverable ID"))
        if not did:
            continue
        deliverables.append(Deliverable(
            id=did,
            title=clean(d.get("Title")),
            deliverable_id=did,
            task_id=clean(d.get("Task ID")),
            description=clean(d.get("Description")),
            status=clean(d.get("Status")),
            start=parse_date(d.get("Start Date")),
            end=parse_date(d.get("End Date")),
            deadline=parse_date(d.get("Deadline")),
            percent_complete=parse_percent(d.get("% Complete")),
            time_allocated=parse_float(d.get("Time Allocated")),
            time_spent=parse_float(d.get("Time Spent")),
        ))
    return deliverables


def _build_hierarchy(projects: list[Project], tasks: list[Task], deliverables: list[Deliverable]) -> None:
    """Link tasks → projects and deliverables → tasks by FK."""
    proj_map = {p.project_id: p for p in projects}
    task_map = {t.task_id: t for t in tasks}

    for t in tasks:
        proj = proj_map.get(t.project_id)
        if proj:
            proj.add_task(t)

    for d in deliverables:
        task = task_map.get(d.task_id)
        if task:
            task.add_deliverable(d)


# ── Public API ─────────────────────────────────────────────────────────────────

def load_profile_from_workbook(
    wb: Workbook,
    *,
    profile_name: str = "",
    company: str = "",
    role: str = "",
    email: str = "",
    phone: str = "",
    recipient_name: str = "",
    recipient_email: str = "",
    workbook_filename: str = "",
    daily_hours_budget: float = 8.0,
) -> Profile:
    """Build a full Profile hierarchy from the 6-sheet workbook schema."""
    profile = Profile(
        id=f"profile:{company or 'default'}",
        title=profile_name or company or "Default",
        company=company,
        role=role,
        email=email,
        phone=phone,
        recipient_name=recipient_name,
        recipient_email=recipient_email,
        workbook_filename=workbook_filename,
        daily_hours_budget=daily_hours_budget,
        status="Active",
    )

    projects = _read_projects(wb)
    tasks = _read_domain_tasks(wb)
    deliverables = _read_deliverables(wb)
    _build_hierarchy(projects, tasks, deliverables)
    for p in projects:
        profile.add_project(p)

    return profile
