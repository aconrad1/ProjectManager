"""Write domain objects to an openpyxl Workbook.

Operates on the 6-sheet schema (Projects / Tasks / Deliverables).
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from helpers.domain.profile import Profile
from helpers.domain.project import Project
from helpers.domain.task import Task
from helpers.domain.deliverable import Deliverable
from helpers.data.tasks import clean
from helpers.schema.sheets import (
    SHEET_PROJECTS, SHEET_TASKS, SHEET_DELIVERABLES,
)
from helpers.schema.columns import (
    PROJECTS_COLUMNS, TASKS_COLUMNS, DELIVERABLES_COLUMNS,
    headers_for, column_index,
)
from helpers.schema.ids import next_project_id, next_task_id, next_deliverable_id


# ── Row writers ─────────────────────────────────────────────────────────────────

def _find_row_by_id(ws, id_col: int, target_id: str) -> int | None:
    """Return the 1-based row index for *target_id* in *id_col* (1-based), or None."""
    for row_idx in range(2, ws.max_row + 1):
        if clean(ws.cell(row=row_idx, column=id_col).value) == target_id:
            return row_idx
    return None


def _write_project_row(ws, row_idx: int, project: Project) -> None:
    """Write a Project domain object into a row on the Projects sheet."""
    values = {
        "Project ID":     project.project_id,
        "Title":          project.title,
        "Category":       project.category,
        "Supervisor":     project.supervisor,
        "Site":           project.site,
        "Description":    project.description,
        "Status":         project.status,
        "Priority":       project.priority,
        "Start Date":     project.start,
        "End Date":       project.end,
        "Deadline":       project.deadline,
        "Date Completed": project.date_completed,
        "Notes":          project.notes,
    }
    for i, col_def in enumerate(PROJECTS_COLUMNS):
        ws.cell(row=row_idx, column=i + 1, value=values.get(col_def.name, ""))


def _write_task_row(ws, row_idx: int, task: Task) -> None:
    """Write a Task domain object into a row on the Tasks sheet."""
    values = {
        "Task ID":            task.task_id,
        "Project ID":         task.project_id,
        "Title":              task.title,
        "Description":        task.description,
        "Supervisor":         task.supervisor,
        "Site":               task.site,
        "Status":             task.status,
        "Priority":           task.priority,
        "Start Date":         task.start,
        "End Date":           task.end,
        "Deadline":           task.deadline,
        "Status Commentary":  task.commentary,
        "Date Completed":     task.date_completed,
        "Scheduled Date":     task.scheduled_date,
    }
    for i, col_def in enumerate(TASKS_COLUMNS):
        ws.cell(row=row_idx, column=i + 1, value=values.get(col_def.name, ""))


def _write_deliverable_row(ws, row_idx: int, deliverable: Deliverable) -> None:
    """Write a Deliverable into a row on the Deliverables sheet."""
    values = {
        "Deliverable ID": deliverable.deliverable_id,
        "Task ID":        deliverable.task_id,
        "Title":          deliverable.title,
        "Description":    deliverable.description,
        "Status":         deliverable.status,
        "Start Date":     deliverable.start,
        "End Date":       deliverable.end,
        "Deadline":       deliverable.deadline,
        "% Complete":     deliverable.percent_complete,
        "Time Allocated": deliverable.time_allocated,
        "Time Spent":     deliverable.time_spent,
    }
    for i, col_def in enumerate(DELIVERABLES_COLUMNS):
        ws.cell(row=row_idx, column=i + 1, value=values.get(col_def.name, ""))


# ── Public add / update / delete ───────────────────────────────────────────────

def add_project_row(wb: Workbook, project: Project) -> str:
    """Append a new project row. Assigns an ID if missing. Returns the project_id."""
    ws = wb[SHEET_PROJECTS]
    if not project.project_id:
        project.project_id = next_project_id(ws)
        project.id = project.project_id
    _write_project_row(ws, ws.max_row + 1, project)
    return project.project_id


def add_task_row(wb: Workbook, task: Task) -> str:
    """Append a new task row. Assigns an ID if missing. Returns the task_id."""
    ws = wb[SHEET_TASKS]
    if not task.task_id:
        task.task_id = next_task_id(ws)
        task.id = task.task_id
    _write_task_row(ws, ws.max_row + 1, task)
    return task.task_id


def add_deliverable_row(wb: Workbook, deliverable: Deliverable) -> str:
    """Append a new deliverable row. Assigns an ID if missing. Returns the deliverable_id."""
    ws = wb[SHEET_DELIVERABLES]
    if not deliverable.deliverable_id:
        deliverable.deliverable_id = next_deliverable_id(ws)
        deliverable.id = deliverable.deliverable_id
    _write_deliverable_row(ws, ws.max_row + 1, deliverable)
    return deliverable.deliverable_id


def update_project_row(wb: Workbook, project: Project) -> bool:
    """Find and overwrite the project row by ID. Returns False if not found."""
    ws = wb[SHEET_PROJECTS]
    row = _find_row_by_id(ws, 1, project.project_id)
    if row is None:
        return False
    _write_project_row(ws, row, project)
    return True


def update_task_row(wb: Workbook, task: Task) -> bool:
    """Find and overwrite the task row by ID. Returns False if not found."""
    ws = wb[SHEET_TASKS]
    row = _find_row_by_id(ws, 1, task.task_id)
    if row is None:
        return False
    _write_task_row(ws, row, task)
    return True


def update_deliverable_row(wb: Workbook, deliverable: Deliverable) -> bool:
    """Find and overwrite the deliverable row by ID. Returns False if not found."""
    ws = wb[SHEET_DELIVERABLES]
    row = _find_row_by_id(ws, 1, deliverable.deliverable_id)
    if row is None:
        return False
    _write_deliverable_row(ws, row, deliverable)
    return True


def delete_row_by_id(wb: Workbook, sheet_name: str, target_id: str) -> bool:
    """Delete a row from *sheet_name* by its ID column. Returns False if not found."""
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]
    row = _find_row_by_id(ws, 1, target_id)
    if row is None:
        return False
    ws.delete_rows(row)
    return True


# ── Full profile sync (new schema) ────────────────────────────────────────────

def save_profile_to_workbook(profile: Profile, wb: Workbook) -> None:
    """Sync the full Profile hierarchy into the workbook.

    Rewrites Projects / Tasks / Deliverables sheets.
    """
    for sheet_name, columns, writer_fn, items in [
        (SHEET_PROJECTS, PROJECTS_COLUMNS, _write_project_row,
         profile.projects),
        (SHEET_TASKS, TASKS_COLUMNS, _write_task_row,
         [t for p in profile.projects for t in p.tasks]),
        (SHEET_DELIVERABLES, DELIVERABLES_COLUMNS, _write_deliverable_row,
         [d for p in profile.projects for t in p.tasks for d in t.deliverables]),
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        for i, item in enumerate(items):
            writer_fn(ws, i + 2, item)
