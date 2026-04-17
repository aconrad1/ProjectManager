"""
pipelines/extract_baseline.py

Extracts all data from the Excel workbook into the star-schema CSV model.

Usage:
    python pipelines/extract_baseline.py "Demand Plan 0409 copy.xlsx"

The workbook is treated as read-only after extraction; never modify it.
"""

import sys
import csv
import uuid
import re
import calendar
import argparse
from datetime import datetime, date, timezone
from pathlib import Path

# Ensure Unicode output works on Windows terminals
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

BASE = Path(__file__).resolve().parent.parent

DIM_PERSON   = BASE / "dimensions" / "dim_person.csv"
DIM_FACILITY = BASE / "dimensions" / "dim_facility.csv"
DIM_PROJECT  = BASE / "dimensions" / "dim_project.csv"
DIM_DATE     = BASE / "dimensions" / "dim_date.csv"
DIM_GROUP    = BASE / "dimensions" / "dim_group.csv"
FACT_HOURS   = BASE / "facts"      / "fact_hours.csv"

LOOKUP_SHEET   = "Lookup"
# Process newest/widest sheet first so it takes precedence on duplicates
INPUT_SHEETS   = ["Input New Projects Load", "Input"]

HEADER_ROW     = 6   # 1-based row with column labels
DATE_ROW       = 5   # 1-based row with month dates per column
DATA_START_ROW = 7   # 1-based first data row

COL_GROUP    = 0
COL_FACILITY = 1
COL_PRIORITY = 2
COL_PROJECT  = 3
DATA_COL_START = 4

SOURCE_TAG = "baseline"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def norm_name(s):
    """Strip whitespace and leading ** markers used for new-hire flags."""
    if s is None:
        return None
    return re.sub(r"^\*+", "", str(s)).strip()


def month_last_day(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


def dt_to_date_id(dt) -> str:
    """datetime -> 'YYYY-MM'"""
    return dt.strftime("%Y-%m")


def fiscal_year_period(d: date):
    """AltaGas fiscal year runs Jan-Dec (same as calendar year)."""
    return d.year, d.month


def new_uuid() -> str:
    return str(uuid.uuid4())


# ---------------------------------------------------------------------------
# Step 1 – Extract people and groups from the Lookup sheet
# ---------------------------------------------------------------------------

def extract_lookup(wb):
    """
    Returns:
        people   : dict  norm_name -> person dict
        groups   : dict  group_name -> group_id  (SAP cost-centre string)
    """
    ws = wb[LOOKUP_SHEET]
    people: dict[str, dict] = {}
    groups: dict[str, str]  = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        _, grp_sap, grp_name, person_sap, person_name = (
            row[0], row[1], row[2], row[3], row[4]
        )
        if person_name is None:
            continue

        pname = norm_name(str(person_name))
        if not pname or pname in ("0",):
            continue

        # Reject placeholder numeric-zero IDs
        try:
            if int(person_sap) == 0:
                continue
        except (TypeError, ValueError):
            pass

        pid = str(person_sap).strip() if person_sap is not None else None
        if not pid:
            continue

        grp = str(grp_name).strip() if grp_name else ""

        parts   = pname.split(",", 1)
        last    = parts[0].strip()
        first   = parts[1].strip() if len(parts) > 1 else ""

        people[pname] = {
            "person_id":      pid,
            "full_name":      pname,
            "first_name":     first,
            "last_name":      last,
            "role":           "",
            "group":          grp,
            "email":          "",
            "active_flag":    "true",
            "effective_start":"",
            "effective_end":  "",
        }

        if grp and grp_sap:
            groups[grp] = str(grp_sap).strip()

    return people, groups


# ---------------------------------------------------------------------------
# Step 2 – Parse a wide-format Input sheet
# ---------------------------------------------------------------------------

def parse_input_sheet(ws, sheet_name: str):
    """
    Returns:
        projects       : dict  project_name -> {facility, project_type, priority}
        hours_records  : list  of {person_norm, project_name, date_id, hours}
        extra_people   : dict  norm_name -> partial person dict (not in Lookup)
        all_date_ids   : set   of all YYYY-MM strings found in column headers
    """
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(all_rows) < HEADER_ROW:
        return {}, [], {}

    date_row   = all_rows[DATE_ROW - 1]
    header_row = all_rows[HEADER_ROW - 1]

    # Build col_index -> (norm_person_name, date_id)
    col_map: dict[int, tuple[str, str]] = {}
    for ci in range(DATA_COL_START, len(header_row)):
        col_hdr  = header_row[ci]
        col_date = date_row[ci]
        if col_hdr is None or col_date is None:
            continue

        col_label = str(col_hdr).strip()
        # Skip "Month Total" summary columns
        if re.search(r"\bTotal\b", col_label, re.IGNORECASE):
            continue

        nname = norm_name(col_label)
        if not nname:
            continue

        if not isinstance(col_date, datetime):
            continue

        col_map[ci] = (nname, dt_to_date_id(col_date))

    all_date_ids: set[str] = {date_id for _, date_id in col_map.values()}

    projects: dict[str, dict]  = {}
    hours_records: list[dict]  = []
    extra_people: dict[str, dict] = {}

    for row in all_rows[DATA_START_ROW - 1:]:
        ptype    = row[COL_GROUP]
        facility = row[COL_FACILITY]
        priority = row[COL_PRIORITY]
        proj_raw = row[COL_PROJECT]

        if proj_raw is None or str(proj_raw).strip() == "":
            continue

        proj_name = str(proj_raw).strip()
        fac       = str(facility).strip() if facility  else ""
        pt        = str(ptype   ).strip() if ptype     else ""
        pri       = str(priority).strip() if priority  else ""

        if proj_name not in projects:
            projects[proj_name] = {
                "project_name": proj_name,
                "facility":     fac,
                "project_type": pt,
                "priority":     pri,
            }

        for ci, (pname_norm, date_id) in col_map.items():
            if ci >= len(row):
                continue
            val = row[ci]
            if val is None:
                continue
            try:
                hrs = float(val)
            except (TypeError, ValueError):
                continue
            if hrs == 0.0:
                continue

            # Track anyone referenced in column headers but absent from Lookup
            if pname_norm not in extra_people:
                extra_people[pname_norm] = {"full_name": pname_norm}

            hours_records.append({
                "person_norm": pname_norm,
                "project_name": proj_name,
                "date_id":      date_id,
                "hours":        hrs,
            })

    return projects, hours_records, extra_people, all_date_ids


# ---------------------------------------------------------------------------
# Step 3 – Build dimension tables
# ---------------------------------------------------------------------------

def build_facilities(all_projects: dict):
    """Return list of facility dicts, keyed by facility_name."""
    fac_map: dict[str, dict] = {}
    for pdata in all_projects.values():
        fname = pdata["facility"]
        if not fname or fname in fac_map:
            continue
        fac_id = "FAC_" + re.sub(r"[^A-Z0-9]+", "_",
                                  fname.upper()).strip("_")
        fac_map[fname] = {
            "facility_id":   fac_id,
            "facility_name": fname,
            "region":        "",
            "business_unit": "",
            "active_flag":   "true",
        }
    return fac_map


def build_projects(all_projects: dict, fac_map: dict):
    """Return list of project dicts."""
    proj_list = []
    seen: set[str] = set()
    for proj_name, pdata in all_projects.items():
        if proj_name in seen:
            continue
        seen.add(proj_name)
        fac_obj = fac_map.get(pdata["facility"], {})
        proj_list.append({
            "project_id":   new_uuid(),
            "project_code": re.sub(r"[^A-Z0-9]+", "_",
                                   proj_name.upper())[:40].strip("_"),
            "project_name": proj_name,
            "project_type": pdata["project_type"],
            "facility_id":  fac_obj.get("facility_id", ""),
            "priority":     pdata["priority"],
            "status":       "Active",
            "active_flag":  "true",
        })
    return proj_list


def build_dates(date_ids: set):
    """Return list of dim_date dicts for every YYYY-MM in date_ids."""
    dates = []
    for did in sorted(date_ids):
        year, month = int(did[:4]), int(did[5:7])
        ms = date(year, month, 1)
        me = month_last_day(year, month)
        fy, fp = fiscal_year_period(ms)
        dates.append({
            "date_id":        did,
            "calendar_month": did,
            "month_start":    ms.isoformat(),
            "month_end":      me.isoformat(),
            "fiscal_year":    fy,
            "fiscal_period":  fp,
        })
    return dates


def build_groups(groups: dict):
    """groups dict: name -> SAP group id."""
    return [
        {
            "group_id":   gid,
            "group_name": gname,
            "manager":    "",
            "active_flag":"true",
        }
        for gname, gid in sorted(groups.items())
    ]


# ---------------------------------------------------------------------------
# Step 4 – Write CSVs
# ---------------------------------------------------------------------------

PERSON_COLS   = ["person_id","full_name","first_name","last_name","role",
                 "group","email","active_flag","effective_start","effective_end"]
FACILITY_COLS = ["facility_id","facility_name","region","business_unit","active_flag"]
PROJECT_COLS  = ["project_id","project_code","project_name","project_type",
                 "facility_id","priority","status","active_flag"]
DATE_COLS     = ["date_id","calendar_month","month_start","month_end",
                 "fiscal_year","fiscal_period"]
GROUP_COLS    = ["group_id","group_name","manager","active_flag"]
FACT_COLS     = ["fact_id","date_id","person_id","project_id","hours",
                 "source","created_at"]


def write_csv(path: Path, fieldnames: list, rows: list):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote {len(rows):>6} rows  ->  {path.relative_to(BASE)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Extract baseline from Excel workbook.")
    parser.add_argument("workbook", help="Path to the Excel workbook")
    args = parser.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.is_absolute():
        wb_path = Path.cwd() / wb_path
    if not wb_path.exists():
        sys.exit(f"Workbook not found: {wb_path}")

    print(f"Loading workbook: {wb_path.name}")
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    # ── People & groups ─────────────────────────────────────────────────────
    print("\nExtracting people from Lookup sheet…")
    people_by_name, groups = extract_lookup(wb)
    print(f"  found {len(people_by_name)} people in Lookup")

    # ── Hours data from Input sheets ────────────────────────────────────────
    all_projects: dict[str, dict] = {}
    all_hours:    list[dict]      = []
    all_extra:    dict[str, dict] = {}
    all_header_dates: set[str]    = set()
    seen_fact_keys: set[tuple]    = set()

    for sheet_name in INPUT_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  sheet '{sheet_name}' not found – skipping")
            continue
        print(f"\nParsing sheet: {sheet_name}")
        ws = wb[sheet_name]
        projects, hours, extras, sheet_dates = parse_input_sheet(ws, sheet_name)
        print(f"  {len(projects)} projects, {len(hours)} non-zero hour cells, "
              f"{len(sheet_dates)} months in headers")

        for pname, pdata in projects.items():
            if pname not in all_projects:
                all_projects[pname] = pdata

        for rec in hours:
            key = (rec["person_norm"], rec["project_name"], rec["date_id"])
            if key not in seen_fact_keys:
                seen_fact_keys.add(key)
                all_hours.append(rec)

        for name, pdata in extras.items():
            if name not in all_extra:
                all_extra[name] = pdata

        all_header_dates.update(sheet_dates)

    wb.close()

    # ── Resolve extra people not in Lookup ──────────────────────────────────
    unknown_count = 0
    for name, _ in all_extra.items():
        if name not in people_by_name:
            pid = "EXT_" + new_uuid()[:8].upper()
            parts  = name.split(",", 1)
            last   = parts[0].strip()
            first  = parts[1].strip() if len(parts) > 1 else ""
            people_by_name[name] = {
                "person_id":      pid,
                "full_name":      name,
                "first_name":     first,
                "last_name":      last,
                "role":           name,   # treat placeholder names as role labels
                "group":          "",
                "email":          "",
                "active_flag":    "true",
                "effective_start":"",
                "effective_end":  "",
            }
            unknown_count += 1

    if unknown_count:
        print(f"\n  {unknown_count} people in column headers not found in Lookup "
              f"-> assigned EXT_* IDs")

    # -- Build dimensions
    print("\nBuilding dimensions...")
    fac_map     = build_facilities(all_projects)
    proj_list   = build_projects(all_projects, fac_map)
    proj_by_name = {p["project_name"]: p for p in proj_list}

    date_ids = all_header_dates | {r["date_id"] for r in all_hours}
    date_list = build_dates(date_ids)
    group_list = build_groups(groups)

    # ── Build fact rows ──────────────────────────────────────────────────────
    print("Building fact_hours…")
    created_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    fact_rows: list[dict] = []
    skipped = 0

    for rec in all_hours:
        person = people_by_name.get(rec["person_norm"])
        project = proj_by_name.get(rec["project_name"])

        if person is None or project is None:
            skipped += 1
            continue

        fact_rows.append({
            "fact_id":    new_uuid(),
            "date_id":    rec["date_id"],
            "person_id":  person["person_id"],
            "project_id": project["project_id"],
            "hours":      rec["hours"],
            "source":     SOURCE_TAG,
            "created_at": created_at,
        })

    if skipped:
        print(f"  WARNING: {skipped} hour records skipped (unresolved person or project)")

    # ── Write all CSVs ───────────────────────────────────────────────────────
    print("\nWriting CSVs…")
    write_csv(DIM_PERSON,   PERSON_COLS,   list(people_by_name.values()))
    write_csv(DIM_FACILITY, FACILITY_COLS, list(fac_map.values()))
    write_csv(DIM_PROJECT,  PROJECT_COLS,  proj_list)
    write_csv(DIM_DATE,     DATE_COLS,     date_list)
    write_csv(DIM_GROUP,    GROUP_COLS,    group_list)
    write_csv(FACT_HOURS,   FACT_COLS,     fact_rows)

    # ── Summary ──────────────────────────────────────────────────────────────
    total_hours = sum(r["hours"] for r in fact_rows)
    print(f"\nBaseline extraction complete")
    print(f"  people:    {len(people_by_name)}")
    print(f"  facilities:{len(fac_map)}")
    print(f"  projects:  {len(proj_list)}")
    print(f"  months:    {len(date_list)}")
    print(f"  groups:    {len(group_list)}")
    print(f"  fact rows: {len(fact_rows)}")
    print(f"  total hrs: {total_hours:,.1f}")
    print(f"\nVerify: sum of fact_hours where source='baseline' should match")
    print(f"        the grand total in the original workbook.")


if __name__ == "__main__":
    main()
